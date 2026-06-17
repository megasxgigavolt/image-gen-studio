"""
Image Gen Studio — NB2 Edition
Visual plan (Excel) → GPT-4o prompt suggestion → Gemini image generation
"""

import json
import re
import os
import sys
import base64
import random
import threading
import time
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from io import BytesIO

import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
from PIL import Image, ImageTk
import openpyxl
from openai import OpenAI
from google.oauth2 import service_account
from google.auth.transport.requests import Request
from google import genai
from google.genai import types
from dotenv import load_dotenv

# ── Paths ─────────────────────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

SA_KEY_FILE    = BASE_DIR / "beneath-the-fins-843aa8608070.json"
GCP_PROJECT_ID = "beneath-the-fins"
STATE_FILE     = BASE_DIR / "generation_state.json"
SETTINGS_FILE  = BASE_DIR / "settings.json"

for _p in [BASE_DIR / ".env", BASE_DIR / "other automations" / ".env"]:
    if _p.exists():
        load_dotenv(_p)
        break

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
GPT_MODEL      = "gpt-4o"

# ── Logging ───────────────────────────────────────────────────────────────────
_LOG_DIR = BASE_DIR / "logs"
_LOG_DIR.mkdir(exist_ok=True)
_logger = logging.getLogger("IGS")
_logger.setLevel(logging.DEBUG)
if not _logger.handlers:
    _fh = RotatingFileHandler(
        str(_LOG_DIR / "image_gen_studio.log"),
        maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8",
    )
    _fh.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)-5s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    ))
    _logger.addHandler(_fh)
    _sh = logging.StreamHandler()
    _sh.setLevel(logging.INFO)
    _sh.setFormatter(logging.Formatter("[IGS %(levelname)-5s] %(message)s"))
    _logger.addHandler(_sh)
_logger.info("=" * 60)
_logger.info("Image Gen Studio — NB2 Edition  starting up")

# ── Theme ─────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

C = {
    "app":          "#FAF5EC",
    "sidebar":      "#EDE3D0",
    "header":       "#1E1208",
    "header_text":  "#F5DEB3",
    "panel":        "#FFFBF4",
    "input":        "#FFF8EE",
    "chat_bg":      "#FEFCF8",
    "prompt_tray":  "#F2E8D5",
    "text":         "#1F120A",
    "text_mid":     "#6B4E35",
    "text_muted":   "#A08060",
    "accent":       "#B5703A",
    "accent_dark":  "#8B4513",
    "chip_bg":      "#EDD9B5",
    "chip_text":    "#5C2E00",
    "divider":      "#DCC9A8",
    "sash":         "#C8B490",
    "btn_blue":     "#1A5FA8",
    "btn_green":    "#27643B",
    "btn_red":      "#A8372A",
    "btn_brown":    "#5D4037",
    "still_todo":   "#EBE4FA",
    "still_todo_t": "#3B1A78",
    "still_done":   "#D5EDDB",
    "still_done_t": "#1A5E28",
    "still_pend":   "#FFF3CD",
    "still_pend_t": "#7B4F00",
    "chat_gpt":     "#0D47A1",
    "chat_user":    "#1B5E20",
    "chat_sys":     "#BF360C",
}

FONT = "Segoe UI"

def F(size: int, weight: str = "bold") -> ctk.CTkFont:
    return ctk.CTkFont(family=FONT, size=size, weight=weight)

def make_combo(parent, values, variable, width=None, **kw) -> ctk.CTkComboBox:
    kwargs = dict(
        values=values, variable=variable,
        fg_color=C["input"], text_color=C["text"],
        border_color=C["divider"], border_width=1,
        button_color=C["accent"], button_hover_color=C["accent_dark"],
        dropdown_fg_color="#FFFBF4", dropdown_text_color=C["text"],
        dropdown_hover_color=C["chip_bg"],
        font=F(14), dropdown_font=F(13),
        state="readonly",
    )
    if width:
        kwargs["width"] = width
    kwargs.update(kw)
    return ctk.CTkComboBox(parent, **kwargs)

# ── Options ───────────────────────────────────────────────────────────────────
ART_STYLES = [
    "Photorealistic", "Cinematic / Film Still", "Documentary",
    "Watercolor", "Oil Painting", "Digital Concept Art",
    "Comic Book / Illustration",
    "Anime / Manga", "Pencil Sketch", "Minimalist / Flat",
    "Vintage / Retro", "Custom (see Extra Notes)",
]
CAMERA_ANGLES = [
    "Wide Shot", "Medium Shot", "Close-Up", "Extreme Close-Up",
    "Bird's Eye View", "Low Angle", "Eye Level",
    "Over the Shoulder", "Dutch Angle", "Custom (see Extra Notes)",
]
MOODS = [
    "Neutral", "Dramatic / Intense", "Serene / Peaceful",
    "Tense / Anxious", "Warm & Cozy", "Cold / Distant",
    "Mysterious", "Cheerful / Upbeat", "Melancholic",
    "Custom (see Extra Notes)",
]
LIGHTING = [
    "Natural Daylight", "Soft Diffused", "Golden Hour",
    "Studio / Clean", "Dramatic Side", "Backlit / Silhouette",
    "Low Key / Dark", "High Key / Bright", "Night / Moonlit",
    "Custom (see Extra Notes)",
]
COLOR_PALETTES = [
    "Natural", "Warm Tones", "Cool Tones", "Muted / Desaturated",
    "High Contrast", "Pastel", "Monochromatic", "Vibrant",
    "Custom (see Extra Notes)",
]
DEPTH_OF_FIELD = [
    "Shallow DoF (bokeh bg)", "Deep DoF (all sharp)",
    "Medium DoF", "Tilt-Shift", "Macro / Extreme Close",
    "Custom (see Extra Notes)",
]

AI_DECIDE = "Let AI Decide"

# GPT model tiers — only multimodal/chat-capable models relevant for prompt & vision work
HC_MODELS = ["gpt-4o", "gpt-4.1", "gpt-5", "gpt-5.1", "o3"]      # Higher Capacity
HV_MODELS = ["gpt-4o-mini", "gpt-4.1-mini", "gpt-4.1-nano",
             "gpt-5-mini", "o4-mini"]                               # Higher Volume
ALL_GPT_MODELS = HC_MODELS + HV_MODELS

_SPINNER = ["|", "/", "-", "\\"]

def _build_extraction_prompt(ref_desc: str) -> tuple[str, str]:
    """Return (system_msg, user_text) for the style-extraction GPT call.

    Produces 8 JSON fields:
      art_style, camera_angle, mood, lighting, color_palette, depth_of_field,
      extra_notes, style_prompt, main_subject
    """
    subject_focus = (
        f"The user specifically wants to extract: '{ref_desc}'. "
        f"For 'main_subject', focus entirely on describing '{ref_desc}' with maximum precision."
        if ref_desc else ""
    )
    system_msg = (
        "You are an elite visual analyst whose output will be fed verbatim into AI image generation prompts. "
        "Extract technical rendering details so precisely that another AI reproduces the EXACT same visual without seeing the original. "
        "Rules: "
        "(1) Copy dropdown option strings EXACTLY — letter for letter. "
        "(2) extra_notes = ONLY rendering/technique details — no subjects, no objects, no scene descriptions. "
        "(3) style_prompt = direct actionable instructions starting with verbs (Use/Apply/Render/Add). "
        "    NEVER use 'capture', 'mirror', 'reflect', 'evoke', 'ensure consistency', or 'cohesive'. "
        "(4) main_subject = a CONDITIONAL appearance rule starting with 'SUBJECT STYLE:' that says "
        "    'When a [type] appears in the scene, render it as:' then lists hex colour regions, "
        "    markings, fur/texture rendering, and ends with 'maintain natural real-world proportions'. "
        "    NEVER use 'always depict' or 'include in every image' — it is conditional on scene context."
    )
    user_text = (
        f"Perform a deep visual style and subject analysis of this image. {subject_focus}\n\n"
        f"Return a JSON object with EXACTLY these keys:\n\n"

        f'"art_style": one of {json.dumps(ART_STYLES)}\n'
        f'"camera_angle": one of {json.dumps(CAMERA_ANGLES)}\n'
        f'"mood": one of {json.dumps(MOODS)}\n'
        f'"lighting": one of {json.dumps(LIGHTING)}\n'
        f'"color_palette": one of {json.dumps(COLOR_PALETTES)}\n'
        f'"depth_of_field": one of {json.dumps(DEPTH_OF_FIELD)}\n\n'

        f'"extra_notes": HIGHLY DETAILED rendering technique only (NO subjects/objects). Must include:\n'
        f'  - Outline weight and colour (e.g. "bold 3px solid black ink outlines")\n'
        f'  - Shading method (e.g. "flat cel-shading with 2-3 tonal steps per surface")\n'
        f'  - Exact dominant scene colours WITH hex codes (e.g. "rust orange #C4622D, cream white #FFF5E6")\n'
        f'  - Shadow technique (e.g. "hard-edged shadows, no ambient occlusion")\n'
        f'  - Texture/pattern details (e.g. "45° crosshatch on shadow areas, plaid fabric patterns")\n'
        f'  - Background complexity (e.g. "fully detailed interior with perspective lines")\n\n'

        f'"main_subject": A CONDITIONAL appearance rule for the primary subject. '
        f'Start with "SUBJECT STYLE: When a [species/type] appears in the scene, render it as:" '
        f'then in ONE compact paragraph:\n'
        f'  - Species/type (e.g. "grey American Shorthair tabby cat")\n'
        f'  - Every colour region with hex codes (e.g. "base fur silver-grey #9A9A9A, '
        f'    bold stripe markings charcoal #2D2D2D, eyes amber-gold #C87941")\n'
        f'  - Distinctive markings (e.g. "M-marking on forehead, tabby swirls on flanks")\n'
        f'  - Fur/texture rendering technique (e.g. "fine parallel curved strokes")\n'
        f'  - End with: "Maintain natural real-world proportions — do not exaggerate size."\n'
        f'  NEVER use "always depict", "include in every image", or unconditional language.\n'
        f'  Example: "SUBJECT STYLE: When a cat appears in the scene, render it as a grey '
        f'  American Shorthair tabby: base fur silver-grey #9A9A9A, bold black tabby stripe '
        f'  markings #2D2D2D, amber-gold eyes #C87941, fine parallel fur strokes. '
        f'  Maintain natural real-world proportions — do not exaggerate size."\n\n'

        f'"style_prompt": 6-8 sentences of DIRECT technical style instructions. Each starts with a verb. '
        f'MUST cover in order: (1) outline style with pixel weight, (2) shading/colouring method, '
        f'(3) exact colour palette with 4-6 hex codes and names, (4) shadow/highlight technique, '
        f'(5) texture and pattern details, (6) background rendering approach, '
        f'(7) lighting quality and direction, (8) subject-specific rendering — '
        f'explicitly state how the MAIN SUBJECT should be rendered (fur strokes, eye style, markings). '
        f'Example for a comic-book cat illustration: '
        f'"Render with bold 3px solid black ink outlines on all shapes. '
        f'Apply flat cel-shading using exactly 3 tonal values: base, mid-shadow, and accent highlight. '
        f'Use this restricted palette: rust orange #C4622D, cream white #FFF5E6, silver-grey #9A9A9A, '
        f'charcoal #2D2D2D, warm tan #C8956C, amber #C87941. '
        f'Cast hard-edged shadows with no gradient softness. '
        f'Add 45-degree crosshatch strokes in shadow regions on fabric and fur. '
        f'Render backgrounds with full interior perspective detail and plaid/tartan textile patterns. '
        f'Light from a warm window source at upper-left with golden ambient fill. '
        f'Depict the cat with fine directional parallel fur strokes, bold black tabby stripe markings, '
        f'and large amber eyes with black slit pupils." '
        f'Do NOT use vague terms.\n\n'

        f"Return ONLY valid JSON with those 8 keys. No markdown fences, no extra text."
    )
    return system_msg, user_text

FIELD_OPTIONS_MAP = {
    "art_style": ART_STYLES,
    "camera":    CAMERA_ANGLES,
    "mood":      MOODS,
    "lighting":  LIGHTING,
    "color":     COLOR_PALETTES,
    "dof":       DEPTH_OF_FIELD,
}

def _gpt_call(client: "OpenAI", model: str, messages: list, max_tokens: int = 900):
    """GPT call with model-appropriate parameters.

    o-series (o3, o4-mini…): use developer role + max_completion_tokens.
    gpt-5* / gpt-4.1*: use max_completion_tokens (max_tokens deprecated).
    Everything else: classic max_tokens.
    """
    is_o    = bool(re.match(r"^o\d", model))
    is_new  = is_o or bool(re.match(r"^gpt-5", model)) or bool(re.match(r"^gpt-4\.1", model))
    fixed_msgs = []
    for msg in messages:
        if is_o and msg["role"] == "system":
            fixed_msgs.append({"role": "developer", "content": msg["content"]})
        else:
            fixed_msgs.append(msg)
    kwargs: dict = {"model": model, "messages": fixed_msgs}
    if is_new:
        kwargs["max_completion_tokens"] = max_tokens
    else:
        kwargs["max_tokens"] = max_tokens
    return client.chat.completions.create(**kwargs)


# ═════════════════════════════════════════════════════════════════════════════
class ImageGenStudio(ctk.CTk):
# ═════════════════════════════════════════════════════════════════════════════

    def __init__(self):
        super().__init__()
        self.title("Image Gen Studio — NB2")
        self.geometry("1520x940")
        self.minsize(1300, 820)
        self.configure(fg_color=C["app"])

        self.visual_plan_path:    Path | None = None
        self.stills:              list[dict]  = []
        self.selected_still:      dict | None = None
        self.chat_history:        list[dict]  = []
        self._chat_log:           list[tuple] = []
        self._still_states:       dict        = {}
        self.current_image_bytes: bytes | None = None
        self.ref_image_b64:       str | None  = None
        self.output_dir:          Path        = BASE_DIR / "generated_images"
        self.gen_state:           dict        = {}
        self._nb2_client                      = None
        self._generating                      = False
        self._spin_idx                        = 0
        self._img_data:       Image.Image | None = None
        self._pending_images: dict            = {}
        self.gpt_model_var:   ctk.StringVar  = ctk.StringVar(value="gpt-4o")

        self._check_startup_state()
        self._build_ui()
        self._load_settings()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── State ─────────────────────────────────────────────────────────────────

    def _check_startup_state(self):
        if STATE_FILE.exists():
            try:
                data      = json.loads(STATE_FILE.read_text(encoding="utf-8"))
                completed = data.get("completed", {})
                if completed:
                    resume = messagebox.askyesno(
                        "Resume Previous Session",
                        f"Found {len(completed)} completed still(s) from a previous run.\n\n"
                        "Resume where you left off?\n(No = clear progress and start fresh)",
                    )
                    self.gen_state = data if resume else {"completed": {}}
                    if not resume:
                        STATE_FILE.write_text(json.dumps(self.gen_state, indent=2))
                else:
                    self.gen_state = data
            except Exception:
                self.gen_state = {"completed": {}}
        else:
            self.gen_state = {"completed": {}}

    def _save_state(self):
        STATE_FILE.write_text(json.dumps(self.gen_state, indent=2), encoding="utf-8")

    # ── Root layout ───────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=0, minsize=308)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build_header()
        self._build_left_panel()
        self._build_right_panel()

    # ── Header ────────────────────────────────────────────────────────────────

    def _build_header(self):
        hdr = ctk.CTkFrame(self, height=66, corner_radius=0, fg_color=C["header"])
        hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        hdr.grid_propagate(False)
        hdr.grid_columnconfigure(2, weight=1)

        # Logo / title
        ctk.CTkLabel(
            hdr, text="Image Gen Studio",
            font=F(21, "bold"), text_color=C["header_text"],
        ).grid(row=0, column=0, padx=22, pady=18)

        self.browse_btn = ctk.CTkButton(
            hdr, text="Browse Visual Plan",
            width=200, height=40, corner_radius=8,
            font=F(14, "bold"),
            fg_color=C["btn_brown"], hover_color="#4E342E",
            command=self._browse_plan,
        )
        self.browse_btn.grid(row=0, column=1, padx=14, pady=14)

        self.plan_label = ctk.CTkLabel(
            hdr, text="No plan loaded — click Browse to start",
            font=F(13), text_color="#9E8060",
        )
        self.plan_label.grid(row=0, column=2, padx=12, sticky="w")

        self.status_lbl = ctk.CTkLabel(
            hdr, text="", font=F(13, "bold"), text_color="#81C784",
        )
        self.status_lbl.grid(row=0, column=3, padx=22)

    # ── Left panel ────────────────────────────────────────────────────────────

    def _build_left_panel(self):
        lp = ctk.CTkFrame(self, corner_radius=0, fg_color=C["sidebar"])
        lp.grid(row=1, column=0, sticky="nsew")
        lp.grid_rowconfigure(1, weight=1)
        lp.grid_columnconfigure(0, weight=1)

        top = ctk.CTkFrame(lp, fg_color="transparent")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=(14, 6))
        top.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(top, text="STILLS", font=F(14, "bold"),
                     text_color=C["accent"]).grid(row=0, column=0, sticky="w")
        self.stills_count_lbl = ctk.CTkLabel(top, text="", font=F(12),
                                              text_color=C["text_muted"])
        self.stills_count_lbl.grid(row=0, column=1, sticky="e")

        self.stills_scroll = ctk.CTkScrollableFrame(
            lp, fg_color="transparent",
            scrollbar_button_color=C["divider"],
            scrollbar_button_hover_color=C["accent"],
        )
        self.stills_scroll.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 4))
        self.stills_scroll.grid_columnconfigure(0, weight=1)

        lp.grid_rowconfigure(2, weight=0)
        bot = ctk.CTkFrame(lp, fg_color="transparent")
        bot.grid(row=2, column=0, sticky="ew", padx=6, pady=(0, 8))
        bot.grid_columnconfigure(0, weight=1)

        self.btn_bulk = ctk.CTkButton(
            bot, text="Bulk Generation Settings",
            height=36, corner_radius=8, font=F(13, "bold"),
            fg_color=C["btn_blue"], hover_color="#0D3D6E",
            command=self._open_bulk_dialog,
        )
        self.btn_bulk.grid(row=0, column=0, sticky="ew")

        self.lbl_bulk_progress = ctk.CTkLabel(
            bot, text="", font=F(11), text_color=C["text_muted"],
        )
        self.lbl_bulk_progress.grid(row=1, column=0, pady=(2, 0))

    # ── Right panel ───────────────────────────────────────────────────────────

    def _build_right_panel(self):
        rp = ctk.CTkFrame(self, corner_radius=0, fg_color=C["app"])
        rp.grid(row=1, column=1, sticky="nsew")
        rp.grid_columnconfigure(0, weight=3)
        rp.grid_columnconfigure(1, weight=2)
        rp.grid_rowconfigure(1, weight=1)

        self._build_info_bar(rp)

        left_col = ctk.CTkFrame(rp, fg_color="transparent")
        left_col.grid(row=1, column=0, sticky="nsew", padx=(10, 5), pady=(6, 10))
        left_col.grid_rowconfigure(1, weight=1)
        left_col.grid_columnconfigure(0, weight=1)
        self._build_settings_panel(left_col)
        self._build_chat_panel(left_col)

        right_col = ctk.CTkFrame(rp, fg_color="transparent")
        right_col.grid(row=1, column=1, sticky="nsew", padx=(5, 10), pady=(6, 10))
        right_col.grid_rowconfigure(1, weight=1)
        right_col.grid_columnconfigure(0, weight=1)
        self._build_reference_panel(right_col)
        self._build_preview_panel(right_col)

    # ── Info bar ──────────────────────────────────────────────────────────────

    def _build_info_bar(self, parent):
        bar = ctk.CTkFrame(parent, fg_color=C["panel"], corner_radius=12,
                            border_width=1, border_color=C["divider"])
        bar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=(10, 0))
        bar.grid_columnconfigure(2, weight=1)

        # Chips row
        chips = ctk.CTkFrame(bar, fg_color="transparent")
        chips.grid(row=0, column=0, columnspan=3, sticky="w", padx=16, pady=(10, 4))

        self._ts_chip = ctk.CTkFrame(chips, fg_color=C["chip_bg"], corner_radius=20)
        self._ts_chip.pack(side="left")
        self.lbl_ts = ctk.CTkLabel(
            self._ts_chip,
            text="Timestamp: —",
            font=F(13, "bold"), text_color=C["chip_text"],
        )
        self.lbl_ts.pack(padx=14, pady=5)

        self._dur_chip = ctk.CTkFrame(chips, fg_color=C["chip_bg"], corner_radius=20)
        self._dur_chip.pack(side="left", padx=(10, 0))
        self.lbl_dur = ctk.CTkLabel(
            self._dur_chip,
            text="Duration: —",
            font=F(13, "bold"), text_color=C["chip_text"],
        )
        self.lbl_dur.pack(padx=14, pady=5)

        # Divider line
        ctk.CTkFrame(bar, height=1, fg_color=C["divider"]).grid(
            row=1, column=0, columnspan=3, sticky="ew", padx=16, pady=0
        )

        # Voiceover
        vo_row = ctk.CTkFrame(bar, fg_color="transparent")
        vo_row.grid(row=2, column=0, columnspan=3, sticky="ew", padx=16, pady=(6, 12))
        vo_row.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(vo_row, text='"', font=F(22, "bold"),
                     text_color=C["accent"]).grid(row=0, column=0, sticky="nw", padx=(0, 8))
        self.lbl_vo = ctk.CTkLabel(
            vo_row,
            text="Select a still from the left panel to begin",
            font=F(14), text_color=C["text"],
            wraplength=820, justify="left", anchor="w",
        )
        self.lbl_vo.grid(row=0, column=1, sticky="ew")

    # ── Image Settings ────────────────────────────────────────────────────────

    def _build_settings_panel(self, parent):
        f = ctk.CTkFrame(parent, fg_color=C["panel"], corner_radius=12,
                          border_width=1, border_color=C["divider"])
        f.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        f.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(f, text="Image Settings", font=F(15, "bold"),
                     text_color=C["accent"]).grid(
            row=0, column=0, columnspan=4, padx=16, pady=(12, 8), sticky="w"
        )

        def lbl(parent, text):
            return ctk.CTkLabel(parent, text=text, font=F(13, "bold"),
                                text_color=C["text_mid"])

        def srow(r, t1, v1, o1, t2, v2, o2):
            lbl(f, t1).grid(row=r, column=0, padx=(16, 6), pady=5, sticky="w")
            make_combo(f, o1, v1).grid(row=r, column=1, padx=4, pady=5, sticky="ew")
            lbl(f, t2).grid(row=r, column=2, padx=(10, 6), pady=5, sticky="w")
            make_combo(f, o2, v2).grid(row=r, column=3, padx=(4, 16), pady=5, sticky="ew")

        self.var_style  = ctk.StringVar(value=ART_STYLES[0])
        self.var_camera = ctk.StringVar(value=CAMERA_ANGLES[0])
        self.var_mood   = ctk.StringVar(value=MOODS[0])
        self.var_light  = ctk.StringVar(value=LIGHTING[0])
        self.var_color  = ctk.StringVar(value=COLOR_PALETTES[0])
        self.var_dof    = ctk.StringVar(value=DEPTH_OF_FIELD[0])

        srow(1, "Art Style:",      self.var_style,  ART_STYLES,    "Camera Angle:",    self.var_camera, CAMERA_ANGLES)
        srow(2, "Mood:",           self.var_mood,   MOODS,         "Lighting:",        self.var_light,  LIGHTING)
        srow(3, "Color Palette:",  self.var_color,  COLOR_PALETTES, "Depth of Field:", self.var_dof,    DEPTH_OF_FIELD)

        # Extra notes
        lbl(f, "Extra Notes:").grid(row=4, column=0, padx=(16, 6), pady=(6, 4), sticky="w")
        self.extra_notes = ctk.CTkEntry(
            f, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1,
            placeholder_text="Additional style notes or custom overrides…",
            placeholder_text_color=C["text_muted"],
            font=F(13), height=36,
        )
        self.extra_notes.grid(row=4, column=1, columnspan=3, padx=(4, 16), pady=(6, 4), sticky="ew")

        # System prompt
        lbl(f, "System Prompt:").grid(row=5, column=0, padx=(16, 6), pady=(4, 4), sticky="nw")
        self.sys_prompt_box = ctk.CTkTextbox(
            f, height=48, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1, font=F(13),
        )
        self.sys_prompt_box.grid(row=5, column=1, columnspan=3, padx=(4, 16), pady=(4, 4), sticky="ew")
        self.sys_prompt_box.insert(
            "0.0",
            "You are an expert visual storyteller crafting image-generation prompts for a YouTube video about cats.",
        )

        # Output folder
        lbl(f, "Save Images To:").grid(row=6, column=0, padx=(16, 6), pady=(4, 12), sticky="w")
        out_row = ctk.CTkFrame(f, fg_color="transparent")
        out_row.grid(row=6, column=1, columnspan=3, padx=(4, 16), pady=(4, 12), sticky="ew")
        out_row.grid_columnconfigure(0, weight=1)

        self.out_dir_label = ctk.CTkLabel(
            out_row, text=str(self.output_dir),
            font=F(12), text_color=C["text_mid"], anchor="w",
        )
        self.out_dir_label.grid(row=0, column=0, sticky="ew")

        ctk.CTkButton(
            out_row, text="Change Folder", width=148, height=32,
            font=F(13, "bold"), fg_color=C["btn_brown"], hover_color="#4E342E",
            corner_radius=8, command=self._pick_output_dir,
        ).grid(row=0, column=1, padx=(10, 0))

        # GPT Model selector
        lbl(f, "GPT Model:").grid(row=7, column=0, padx=(16, 6), pady=(4, 12), sticky="w")
        gm_row = ctk.CTkFrame(f, fg_color="transparent")
        gm_row.grid(row=7, column=1, columnspan=3, padx=(4, 16), pady=(4, 12), sticky="ew")
        gm_row.grid_columnconfigure(0, weight=1)
        gm_combo = make_combo(gm_row, ALL_GPT_MODELS, self.gpt_model_var)
        gm_combo.grid(row=0, column=0, sticky="ew")
        gm_combo.configure(command=lambda v: self._update_tier_lbl(v, self._main_tier_lbl))
        self._main_tier_lbl = ctk.CTkLabel(
            gm_row, text="● Higher Capacity", font=F(12), text_color="#27643B")
        self._main_tier_lbl.grid(row=0, column=1, padx=(10, 0))

    # ── Chat panel with resizable panes ───────────────────────────────────────

    def _build_chat_panel(self, parent):
        outer = ctk.CTkFrame(parent, fg_color=C["panel"], corner_radius=12,
                              border_width=1, border_color=C["divider"])
        outer.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
        outer.grid_rowconfigure(1, weight=1)
        outer.grid_columnconfigure(0, weight=1)

        # ── Header row ────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(outer, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 6))
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text="Prompt Chat", font=F(15, "bold"),
                     text_color=C["accent"]).grid(row=0, column=0, sticky="w")

        ctk.CTkButton(
            hdr, text="Reset Chat", width=120, height=32, corner_radius=8,
            font=F(13, "bold"), fg_color=C["btn_red"], hover_color="#8B1A1A",
            command=self._reset_chat,
        ).grid(row=0, column=1)

        # ── Resizable PanedWindow: chat display / prompt editor ───────────────
        pw = tk.PanedWindow(
            outer, orient=tk.VERTICAL,
            sashwidth=7, sashrelief="flat",
            background=C["sash"], bd=0,
            opaqueresize=True,
        )
        pw.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 6))

        # Chat display pane
        chat_outer = tk.Frame(pw, bg=C["chat_bg"])
        self.chat_display = ctk.CTkTextbox(
            chat_outer, fg_color=C["chat_bg"], text_color=C["text"],
            font=F(13), wrap="word",
            border_width=0,
        )
        self.chat_display.pack(fill="both", expand=True, padx=6, pady=6)
        self.chat_display.configure(state="disabled")
        self.chat_display.tag_config("gpt",  foreground=C["chat_gpt"])
        self.chat_display.tag_config("user", foreground=C["chat_user"])
        self.chat_display.tag_config("sys",  foreground=C["chat_sys"])
        pw.add(chat_outer, minsize=120, stretch="always")

        # Prompt editor pane
        tray_outer = tk.Frame(pw, bg=C["prompt_tray"])
        self._build_prompt_tray(tray_outer)
        pw.add(tray_outer, minsize=200, stretch="never")

        # ── Chat input ────────────────────────────────────────────────────────
        ci = ctk.CTkFrame(outer, fg_color="transparent")
        ci.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 12))
        ci.grid_columnconfigure(0, weight=1)

        self.chat_input = ctk.CTkEntry(
            ci, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1,
            placeholder_text="Describe changes, e.g. 'add morning mist, warmer light'…",
            placeholder_text_color=C["text_muted"],
            height=40, font=F(14),
        )
        self.chat_input.grid(row=0, column=0, sticky="ew")
        self.chat_input.bind("<Return>", lambda _: self._send_chat())

        ctk.CTkButton(
            ci, text="Send", width=90, height=40, corner_radius=8,
            font=F(14, "bold"), fg_color=C["btn_brown"], hover_color="#4E342E",
            command=self._send_chat,
        ).grid(row=0, column=1, padx=(8, 0))

    def _build_prompt_tray(self, parent):
        """Inner layout of the resizable prompt editor pane."""
        parent_ctk = ctk.CTkFrame(parent, fg_color=C["prompt_tray"], corner_radius=0)
        parent_ctk.pack(fill="both", expand=True)
        parent_ctk.grid_columnconfigure(0, weight=1)
        parent_ctk.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(
            parent_ctk,
            text="Current Prompt  —  edit freely before generating",
            font=F(12, "bold"), text_color=C["text_mid"],
        ).grid(row=0, column=0, columnspan=2, padx=12, pady=(10, 4), sticky="w")

        self.prompt_editor = ctk.CTkTextbox(
            parent_ctk, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1,
            font=F(14), wrap="word",
        )
        self.prompt_editor.grid(row=1, column=0, columnspan=2, sticky="nsew",
                                 padx=12, pady=(0, 8))

        btn_row = ctk.CTkFrame(parent_ctk, fg_color="transparent")
        btn_row.grid(row=2, column=0, columnspan=2, sticky="ew", padx=12, pady=(0, 12))

        self.btn_suggest = ctk.CTkButton(
            btn_row, text="Suggest Prompt",
            width=168, height=42, corner_radius=8,
            font=F(14, "bold"), fg_color=C["btn_blue"], hover_color="#0D47A1",
            command=self._request_suggestion,
        )
        self.btn_suggest.pack(side="left")

        self.btn_generate = ctk.CTkButton(
            btn_row, text="Generate Image",
            width=168, height=42, corner_radius=8,
            font=F(14, "bold"), fg_color=C["btn_green"], hover_color="#1B4D2E",
            command=self._generate_image,
        )
        self.btn_generate.pack(side="left", padx=(10, 0))

    # ── Reference image panel ─────────────────────────────────────────────────

    def _build_reference_panel(self, parent):
        f = ctk.CTkFrame(parent, fg_color=C["panel"], corner_radius=12,
                          border_width=1, border_color=C["divider"])
        f.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        f.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(f, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 8))
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text="Reference Image", font=F(15, "bold"),
                     text_color=C["accent"]).grid(row=0, column=0, sticky="w")

        ctk.CTkButton(
            hdr, text="Upload", width=88, height=32, corner_radius=8,
            font=F(13, "bold"), fg_color=C["btn_blue"], hover_color="#0D47A1",
            command=self._upload_reference,
        ).grid(row=0, column=1)

        ctk.CTkButton(
            hdr, text="Clear", width=80, height=32, corner_radius=8,
            font=F(13, "bold"), fg_color="#9E9E9E", hover_color="#757575",
            command=self._clear_reference,
        ).grid(row=0, column=2, padx=(6, 0))

        self.ref_label = ctk.CTkLabel(
            f,
            text="No reference image\nUpload to send visual context to GPT-4o",
            text_color=C["text_muted"], height=140,
            font=F(13),
        )
        self.ref_label.grid(row=1, column=0, padx=16, pady=(0, 6), sticky="ew")

        ctk.CTkLabel(f, text="What to pick from this reference image:",
                     font=F(12), text_color=C["text_mid"]).grid(
            row=2, column=0, padx=16, pady=(0, 2), sticky="w")

        desc_row = ctk.CTkFrame(f, fg_color="transparent")
        desc_row.grid(row=3, column=0, padx=16, pady=(0, 14), sticky="ew")
        desc_row.grid_columnconfigure(0, weight=1)

        self.ref_desc_entry = ctk.CTkEntry(
            desc_row, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1, font=F(13), height=34,
            placeholder_text="e.g. color grading, lighting style, mood, framing…",
        )
        self.ref_desc_entry.grid(row=0, column=0, sticky="ew")

        self.btn_extract = ctk.CTkButton(
            desc_row, text="Extract Settings",
            height=34, corner_radius=8, font=F(13, "bold"),
            fg_color=C["btn_red"], hover_color="#8B1A1A",
            command=self._extract_ref_settings,
        )
        self.btn_extract.grid(row=0, column=1, padx=(8, 0))

    # ── Generated image preview ───────────────────────────────────────────────

    def _build_preview_panel(self, parent):
        f = ctk.CTkFrame(parent, fg_color=C["panel"], corner_radius=12,
                          border_width=1, border_color=C["divider"])
        f.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
        f.grid_rowconfigure(1, weight=1)
        f.grid_columnconfigure(0, weight=1)
        self._gen_frame = f

        ctk.CTkLabel(f, text="Generated Image", font=F(15, "bold"),
                     text_color=C["accent"]).grid(row=0, column=0, padx=16, pady=(12, 6), sticky="w")

        # Plain tk.Label — avoids CTkLabel's text/image mode confusion with raw PhotoImages
        self.gen_preview = tk.Label(
            f, text="Image will appear here after generation",
            fg=C["text_muted"], bg=C["panel"],
            font=("Segoe UI", 13),
            compound="center",
        )
        self.gen_preview.grid(row=1, column=0, sticky="nsew", padx=10, pady=6)

        self.btn_approve = ctk.CTkButton(
            f, text="Approve & Save",
            height=46, corner_radius=10,
            font=F(16, "bold"), fg_color=C["btn_green"], hover_color="#1B4D2E",
            command=self._approve_image, state="disabled",
        )
        self.btn_approve.grid(row=2, column=0, padx=24, pady=(6, 20), sticky="ew")

    # ── Plan loading ──────────────────────────────────────────────────────────

    def _browse_plan(self):
        path = filedialog.askopenfilename(
            title="Select Visual Plan",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.visual_plan_path = Path(path)
            self._load_plan()

    def _load_plan(self):
        try:
            wb      = openpyxl.load_workbook(self.visual_plan_path)
            ws      = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]

            def _col(kw):
                return next((i for i, h in enumerate(headers) if kw in h), None)

            ci_start, ci_end = _col("start"), _col("end")
            ci_dur, ci_sent  = _col("duration"), _col("sentence")
            ci_type          = _col("type")

            if any(c is None for c in (ci_start, ci_end, ci_dur, ci_sent, ci_type)):
                messagebox.showerror(
                    "Column Error",
                    "Required columns not found.\n"
                    "Expected: Start Timestamp, End Timestamp, Duration, Sentences, Type",
                )
                return

            self.stills = []
            idx = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[ci_type] or "").strip().lower() == "still":
                    self.stills.append({
                        "index":     idx,
                        "still_id":  f"s{idx}",
                        "start":     str(row[ci_start] or ""),
                        "end":       str(row[ci_end]   or ""),
                        "duration":  row[ci_dur],
                        "voiceover": str(row[ci_sent]  or ""),
                    })
                    idx += 1

            self.plan_label.configure(
                text=f"  {self.visual_plan_path.name}  —  {len(self.stills)} stills"
            )
            self._populate_stills_list()
        except Exception as exc:
            messagebox.showerror("Load Error", f"Failed to load visual plan:\n{exc}")

    def _populate_stills_list(self):
        for w in self.stills_scroll.winfo_children():
            w.destroy()

        completed  = self.gen_state.get("completed", {})
        done_count = sum(1 for s in self.stills if s["still_id"] in completed)
        pend_count = sum(1 for s in self.stills if s["still_id"] in self._pending_images)
        count_txt  = f"{done_count}/{len(self.stills)} done"
        if pend_count:
            count_txt += f"  |  {pend_count} pending"
        self.stills_count_lbl.configure(text=count_txt)

        selected_id = self.selected_still["still_id"] if self.selected_still else None

        for still in self.stills:
            sid  = still["still_id"]
            done = sid in completed
            pend = sid in self._pending_images
            is_sel = (sid == selected_id)
            vo   = still["voiceover"]
            preview = vo[:50] + "…" if len(vo) > 50 else vo

            icon    = "✓ " if done else ("* " if pend else "  ")
            fg      = C["still_done"] if done else (C["still_pend"] if pend else C["still_todo"])
            fg_t    = C["still_done_t"] if done else (C["still_pend_t"] if pend else C["still_todo_t"])
            hover   = "#C0DFC7" if done else ("#FFE57F" if pend else "#D8D0F5")
            btn = ctk.CTkButton(
                self.stills_scroll,
                text=f"{icon}{sid.upper()}   {still['start'][:8]}\n{preview}",
                fg_color=fg, text_color=fg_t, hover_color=hover,
                anchor="w", font=F(12), height=58, corner_radius=8,
                border_width=2 if is_sel else 0,
                border_color=C["accent"] if is_sel else C["divider"],
                command=lambda s=still: self._select_still(s),
            )
            btn.grid(sticky="ew", padx=6, pady=3)

        if pend_count:
            self.lbl_bulk_progress.configure(text=f"{pend_count} pending approval")
        else:
            self.lbl_bulk_progress.configure(text="")

    # ── Still selection ───────────────────────────────────────────────────────

    def _select_still(self, still: dict):
        _logger.info(f"[STILL] Selected: {still['still_id']} | vo={str(still.get('voiceover',''))[:80]!r}")
        if self.selected_still:
            old = self.selected_still["still_id"]
            saved = {
                "history": list(self.chat_history),
                "log":     list(self._chat_log),
                "prompt":  self.prompt_editor.get("0.0", "end"),
            }
            # Preserve bulk_settings so settings are restored when returning to this still
            prev = self._still_states.get(old, {})
            if "bulk_settings" in prev:
                saved["bulk_settings"] = prev["bulk_settings"]
            self._still_states[old] = saved

        self.selected_still = still
        new_id = still["still_id"]

        self.lbl_ts.configure(text=f"{still['start'][:12]}  →  {still['end'][:12]}")
        self.lbl_dur.configure(text=f"⏱  {still['duration']}s on screen")
        self.lbl_vo.configure(text=still["voiceover"])
        self.status_lbl.configure(text=f"  {still['still_id'].upper()}")

        self._img_data = None
        self.current_image_bytes = None
        self._clear_preview()
        self.btn_approve.configure(state="disabled")

        if new_id in self._pending_images:
            self.current_image_bytes = self._pending_images[new_id]
            self._img_data = Image.open(BytesIO(self.current_image_bytes))
            self.btn_approve.configure(state="normal")
            self.after_idle(self._display_gen_image)

        if new_id not in self._still_states:
            self.chat_history = []
            self._chat_log    = []
            self.chat_display.configure(state="normal")
            self.chat_display.delete("0.0", "end")
            self.chat_display.configure(state="disabled")
            self.prompt_editor.delete("0.0", "end")
            # Skip auto-suggestion if already have a pending generated image
            if new_id not in self._pending_images:
                self._request_suggestion(auto=True)
        else:
            state = self._still_states[new_id]
            self.chat_history = list(state["history"])
            self._chat_log    = list(state["log"])
            self.prompt_editor.delete("0.0", "end")
            self.prompt_editor.insert("0.0", state["prompt"])
            # Restore settings if this still was bulk-generated
            if "bulk_settings" in state:
                self._apply_bulk_settings(state["bulk_settings"])
            self.chat_display.configure(state="normal")
            self.chat_display.delete("0.0", "end")
            for tag, label, body in self._chat_log:
                self.chat_display.insert("end", f"\n{label}\n", tag)
                self.chat_display.insert("end", f"{body}\n")
            self.chat_display.see("end")
            self.chat_display.configure(state="disabled")

    # ── Output folder ─────────────────────────────────────────────────────────

    def _pick_output_dir(self):
        chosen = filedialog.askdirectory(title="Choose folder to save generated images")
        if chosen:
            self.output_dir = Path(chosen)
            self.out_dir_label.configure(text=str(self.output_dir))

    # ── Bulk settings sync ────────────────────────────────────────────────────

    def _apply_bulk_settings(self, bs: dict):
        """Apply bulk-generation settings to the main frontend dropdowns/entries."""
        _logger.info(
            f"[SETTINGS→UI] art={bs.get('art_style') or '—'} | "
            f"cam={bs.get('camera') or '—'} | mood={bs.get('mood') or '—'} | "
            f"light={bs.get('lighting') or '—'} | color={bs.get('color') or '—'} | "
            f"dof={bs.get('dof') or '—'}"
        )
        if bs.get("art_style"):  self.var_style.set(bs["art_style"])
        if bs.get("camera"):     self.var_camera.set(bs["camera"])
        if bs.get("mood"):       self.var_mood.set(bs["mood"])
        if bs.get("lighting"):   self.var_light.set(bs["lighting"])
        if bs.get("color"):      self.var_color.set(bs["color"])
        if bs.get("dof"):        self.var_dof.set(bs["dof"])
        notes = bs.get("extra_notes", "")
        self.extra_notes.delete(0, "end")
        if notes:
            self.extra_notes.insert(0, notes)
        sys_p = bs.get("system_prompt", "")
        if sys_p:
            self.sys_prompt_box.delete("0.0", "end")
            self.sys_prompt_box.insert("0.0", sys_p)

    # ── GPT model tier helper ─────────────────────────────────────────────────

    @staticmethod
    def _update_tier_lbl(model: str, lbl: ctk.CTkLabel):
        if model in HC_MODELS:
            lbl.configure(text="● Higher Capacity", text_color="#27643B")
        else:
            lbl.configure(text="● Higher Volume",   text_color="#B8860B")

    # ── Reference image settings extraction ───────────────────────────────────

    def _extract_ref_settings(self):
        if not self.ref_image_b64:
            messagebox.showwarning("No Reference Image",
                                   "Upload a reference image first.")
            return
        if not OPENAI_API_KEY:
            messagebox.showerror("API Key Missing",
                                  "OPENAI_API_KEY not found in .env file.")
            return
        self.btn_extract.configure(state="disabled", text="Extracting…")
        ref_b64  = self.ref_image_b64
        ref_desc = self.ref_desc_entry.get().strip()
        model    = self.gpt_model_var.get()
        threading.Thread(target=self._extract_settings_worker,
                          args=(ref_b64, ref_desc, model), daemon=True).start()

    def _extract_settings_worker(self, ref_b64: str, ref_desc: str, model: str):
        _logger.info(f"[EXTRACT] Start — model={model} | ref_desc={ref_desc!r}")
        try:
            client = OpenAI(api_key=OPENAI_API_KEY)
            sys_msg, user_text = _build_extraction_prompt(ref_desc)
            msgs = [
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": [
                    {"type": "image_url",
                     "image_url": {"url": f"data:image/png;base64,{ref_b64}",
                                   "detail": "high"}},
                    {"type": "text", "text": user_text},
                ]},
            ]
            resp = _gpt_call(client, model, msgs, 1200)
            raw = resp.choices[0].message.content.strip()
            _logger.debug(f"[EXTRACT] Raw GPT response:\n{raw}")
            m   = re.search(r"\{.*\}", raw, re.DOTALL)
            data = json.loads(m.group() if m else raw)
            _logger.info(
                f"[EXTRACT] Parsed — art={data.get('art_style')} | "
                f"cam={data.get('camera_angle')} | mood={data.get('mood')} | "
                f"light={data.get('lighting')} | color={data.get('color_palette')} | "
                f"dof={data.get('depth_of_field')}"
            )
            _logger.info(f"[EXTRACT] Extra notes: {str(data.get('extra_notes',''))[:300]}")
            _logger.info(f"[EXTRACT] Style prompt: {str(data.get('style_prompt',''))[:300]}")
            self.after(0, lambda: self._apply_extracted_settings(data))
        except Exception as exc:
            _logger.error(f"[EXTRACT] Error: {exc}")
            self.after(0, lambda e=str(exc): (
                messagebox.showerror("Extraction Error", e),
                self.btn_extract.configure(state="normal", text="Extract Settings"),
            ))

    def _apply_extracted_settings(self, data: dict):
        def try_set(var, key, options):
            val = data.get(key, "")
            if val in options:
                var.set(val)

        try_set(self.var_style,  "art_style",      ART_STYLES)
        try_set(self.var_camera, "camera_angle",   CAMERA_ANGLES)
        try_set(self.var_mood,   "mood",           MOODS)
        try_set(self.var_light,  "lighting",       LIGHTING)
        try_set(self.var_color,  "color_palette",  COLOR_PALETTES)
        try_set(self.var_dof,    "depth_of_field", DEPTH_OF_FIELD)

        notes = data.get("extra_notes", "").strip()
        if notes:
            self.extra_notes.delete(0, "end")
            self.extra_notes.insert(0, notes)

        style_p = data.get("style_prompt", "").strip()
        subject = data.get("main_subject", "").strip()

        # Combine style directive + subject lock into one system prompt
        # so every Gemini prompt and every per-still GPT call sees both
        full_sys = style_p
        if subject:
            full_sys = f"{style_p}\n\n{subject}" if style_p else subject

        if full_sys:
            self.sys_prompt_box.delete("0.0", "end")
            self.sys_prompt_box.insert("0.0", full_sys)

        _logger.info(
            f"[EXTRACT→UI] Applied — style_prompt_len={len(style_p)} | "
            f"main_subject={subject[:120]!r}"
        )
        self.btn_extract.configure(state="normal", text="Extract Settings")
        messagebox.showinfo(
            "Settings Extracted",
            "Dropdowns, Extra Notes, and System Prompt updated.\n"
            + (f"Subject lock included: {subject[:80]}…" if subject else ""),
        )

    # ── Reference image ───────────────────────────────────────────────────────

    def _upload_reference(self):
        path = filedialog.askopenfilename(
            title="Select Reference Image",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.webp"), ("All", "*.*")],
        )
        if not path:
            return
        try:
            img   = Image.open(path)
            thumb = img.copy()
            thumb.thumbnail((308, 192), Image.LANCZOS)
            photo = ImageTk.PhotoImage(thumb)
            self.ref_label.configure(image=photo, text="")
            self.ref_label._ref = photo
            buf = BytesIO()
            img.save(buf, format="PNG")
            self.ref_image_b64 = base64.b64encode(buf.getvalue()).decode()
        except Exception as exc:
            messagebox.showerror("Image Error", str(exc))

    def _clear_reference(self):
        self.ref_image_b64 = None
        self.ref_label.configure(
            image=None,
            text="No reference image\nUpload to send visual context to GPT-4o",
        )

    # ── Settings summary ──────────────────────────────────────────────────────

    def _settings_block(self) -> str:
        lines = [
            f"Art Style:       {self.var_style.get()}",
            f"Camera Angle:    {self.var_camera.get()}",
            f"Mood:            {self.var_mood.get()}",
            f"Lighting:        {self.var_light.get()}",
            f"Color Palette:   {self.var_color.get()}",
            f"Depth of Field:  {self.var_dof.get()}",
        ]
        notes = self.extra_notes.get().strip()
        if notes:
            lines.append(f"Extra Notes:     {notes}")
        return "\n".join(lines)

    # ── GPT-4o suggestion ─────────────────────────────────────────────────────

    def _request_suggestion(self, auto: bool = False):
        if not self.selected_still:
            if not auto:
                messagebox.showwarning("No Still Selected",
                                       "Select a still from the left panel first.")
            return
        if not OPENAI_API_KEY:
            if not auto:
                messagebox.showerror("API Key Missing",
                                      "OPENAI_API_KEY not found in .env file.")
            return
        self.btn_suggest.configure(state="disabled", text="⏳  Thinking…")
        # Snapshot widget values NOW (main thread) — safe to pass to background thread
        sys_txt  = self.sys_prompt_box.get("0.0", "end").strip()
        settings = self._settings_block()
        ref_b64  = self.ref_image_b64
        ref_desc = self.ref_desc_entry.get().strip()
        still    = self.selected_still
        history  = list(self.chat_history)
        model    = self.gpt_model_var.get()
        threading.Thread(
            target=self._suggestion_worker,
            args=(still, sys_txt, settings, ref_b64, ref_desc, history, model),
            daemon=True,
        ).start()

    def _suggestion_worker(self, still, sys_txt, settings, ref_b64, ref_desc, history, model="gpt-4o"):
        try:
            client = OpenAI(api_key=OPENAI_API_KEY)
            parts: list = []

            if ref_b64:
                parts.append({"type": "image_url",
                               "image_url": {"url": f"data:image/png;base64,{ref_b64}"}})
                guidance = ref_desc if ref_desc else "visual style and aesthetic"
                parts.append({"type": "text",
                               "text": f"The above image is a reference. From it, use: {guidance}"})

            parts.append({"type": "text", "text": (
                f"Write an image-generation prompt for this still:\n\n"
                f"VOICEOVER:\n\"{still['voiceover']}\"\n\n"
                f"STILL INFO:\n"
                f"  Timestamp : {still['start']} → {still['end']}\n"
                f"  On-screen : {still['duration']} seconds\n\n"
                f"IMAGE SETTINGS:\n{settings}\n\n"
                f"Rules:\n"
                f"  • VISUALLY INTERPRET the voiceover — describe what the viewer SEES, "
                f"not what is being narrated. Derive a scene from it; do NOT copy it verbatim.\n"
                f"  • Explicitly mention EVERY setting in the prompt naturally "
                f"(camera angle, art style, mood, lighting, color palette, depth of field).\n"
                f"  • Compose for 16:9 aspect ratio\n"
                f"  • Be specific about subject, composition, lighting, atmosphere\n"
                f"  • NO text, words, or captions inside the image\n\n"
                f"Respond with ONLY the prompt text — no preamble, no explanations."
            )})

            messages = [{"role": "system", "content": sys_txt}]
            messages.extend(history)
            messages.append({"role": "user", "content": parts})

            resp = _gpt_call(client, model, messages, 900)
            suggested = resp.choices[0].message.content.strip()

            new_user = {"role": "user",
                "content": f"[Suggest for {still['still_id']}: '{still['voiceover'][:60]}…']"}
            new_asst = {"role": "assistant", "content": suggested}
            self.after(0, lambda: self._apply_suggestion(suggested, new_user, new_asst))
        except Exception as exc:
            self.after(0, lambda: self._gpt_error(str(exc)))

    def _apply_suggestion(self, text: str, user_msg: dict, asst_msg: dict):
        self.chat_history.append(user_msg)
        self.chat_history.append(asst_msg)
        self._chat_append("GPT-4o suggested prompt:", text, "gpt")
        self.prompt_editor.delete("0.0", "end")
        self.prompt_editor.insert("0.0", text)
        self.btn_suggest.configure(state="normal", text="✦  Suggest Prompt")

    # ── Chat ─────────────────────────────────────────────────────────────────

    def _send_chat(self):
        msg = self.chat_input.get().strip()
        if not msg:
            return
        self.chat_input.delete(0, "end")
        self._chat_append("You:", msg, "user")
        self.btn_suggest.configure(state="disabled", text="⏳  Thinking…")
        # Snapshot all widget values in main thread before handing to background
        sys_txt  = self.sys_prompt_box.get("0.0", "end").strip()
        cur_p    = self.prompt_editor.get("0.0", "end").strip()
        settings = self._settings_block()
        history  = list(self.chat_history)
        model    = self.gpt_model_var.get()
        threading.Thread(
            target=self._chat_worker,
            args=(msg, sys_txt, cur_p, settings, history, model),
            daemon=True,
        ).start()

    def _chat_worker(self, user_msg: str, sys_txt: str, cur_p: str, settings: str,
                     history: list, model: str = "gpt-4o"):
        try:
            client = OpenAI(api_key=OPENAI_API_KEY)
            req = (f"Current prompt:\n\"{cur_p}\"\n\n"
                   f"Image settings:\n{settings}\n\n"
                   f"User feedback: {user_msg}\n\n"
                   f"Return ONLY an updated prompt incorporating this feedback.")
            user_entry = {"role": "user", "content": req}

            messages = [{"role": "system", "content": sys_txt}]
            messages.extend(history)
            messages.append(user_entry)

            resp  = _gpt_call(client, model, messages, 900)
            reply = resp.choices[0].message.content.strip()
            asst_entry = {"role": "assistant", "content": reply}
            self.after(0, lambda: self._apply_refinement(reply, user_entry, asst_entry))
        except Exception as exc:
            self.after(0, lambda: self._gpt_error(str(exc)))

    def _apply_refinement(self, text: str, user_entry: dict, asst_entry: dict):
        self.chat_history.append(user_entry)
        self.chat_history.append(asst_entry)
        self._chat_append("GPT-4o refined prompt:", text, "gpt")
        self.prompt_editor.delete("0.0", "end")
        self.prompt_editor.insert("0.0", text)
        self.btn_suggest.configure(state="normal", text="✦  Suggest Prompt")

    def _gpt_error(self, msg: str):
        self.btn_suggest.configure(state="normal", text="✦  Suggest Prompt")
        messagebox.showerror("GPT-4o Error", msg)

    def _reset_chat(self):
        if messagebox.askyesno("Reset Chat",
                               "Clear chat history for this still and start fresh?"):
            self.chat_history = []
            self._chat_log    = []
            self.chat_display.configure(state="normal")
            self.chat_display.delete("0.0", "end")
            self.chat_display.configure(state="disabled")
            self.prompt_editor.delete("0.0", "end")
            if self.selected_still:
                self._still_states.pop(self.selected_still["still_id"], None)

    def _chat_append(self, label: str, body: str, tag: str = "gpt"):
        self._chat_log.append((tag, label, body))
        self.chat_display.configure(state="normal")
        self.chat_display.insert("end", f"\n{label}\n", tag)
        self.chat_display.insert("end", f"{body}\n")
        self.chat_display.see("end")
        self.chat_display.configure(state="disabled")

    # ── Spinner ───────────────────────────────────────────────────────────────

    def _start_spinner(self):
        self._spin_idx = 0
        self._tick_spinner()

    def _tick_spinner(self):
        if not self._generating:
            return
        self._spin_idx = (self._spin_idx + 1) % len(_SPINNER)
        self.gen_preview.configure(
            image="",
            text=(f"\n\n{_SPINNER[self._spin_idx]}\n\n"
                  "Generating your image…\n"
                  "This may take 15 – 30 seconds"),
            font=("Segoe UI", 17, "bold"),
            fg=C["accent"],
        )
        self.after(80, self._tick_spinner)

    # ── Gemini generation ─────────────────────────────────────────────────────

    def _init_nb2(self):
        if self._nb2_client:
            return self._nb2_client
        if not SA_KEY_FILE.exists():
            raise FileNotFoundError(
                f"Service account key not found:\n{SA_KEY_FILE}\n\n"
                "Place beneath-the-fins-843aa8608070.json next to the application."
            )
        creds = service_account.Credentials.from_service_account_file(
            str(SA_KEY_FILE),
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
        creds.refresh(Request())
        self._nb2_client = genai.Client(
            vertexai=True, project=GCP_PROJECT_ID,
            location="global", credentials=creds,
        )
        return self._nb2_client

    def _generate_image(self):
        if not self.selected_still:
            messagebox.showwarning("No Still", "Select a still first.")
            return
        prompt = self.prompt_editor.get("0.0", "end").strip()
        if not prompt:
            messagebox.showwarning("Empty Prompt",
                                    "Write or generate a prompt before generating.")
            return
        if self._generating:
            return

        # Build Gemini prompt: style directive + settings footer + scene prompt
        sys_p    = self.sys_prompt_box.get("0.0", "end").strip()
        settings = self._settings_block()   # art style, camera, mood, lighting, colour, dof
        parts    = []
        if sys_p:
            parts.append(sys_p)
        parts.append(prompt)
        if settings:
            parts.append(f"APPLY THESE STYLE SETTINGS EXACTLY:\n{settings}")
        gemini_prompt = "\n\n".join(parts)

        sid = self.selected_still["still_id"]
        _logger.info(
            f"[SINGLE-GEN] Start — still={sid} | "
            f"sys_p_len={len(sys_p)} | prompt_len={len(prompt)} | "
            f"gemini_total_len={len(gemini_prompt)}"
        )
        _logger.debug(f"[SINGLE-GEN] Full Gemini prompt:\n{gemini_prompt}")

        self._generating = True
        self._img_data = None
        self._pending_images.pop(sid, None)
        self.gen_preview.configure(image="", text="")
        if hasattr(self.gen_preview, "_ref"):
            self.gen_preview._ref = None
        self.btn_generate.configure(state="disabled", text="▶  Generating…")
        self.btn_approve.configure(state="disabled")
        self._start_spinner()
        threading.Thread(target=self._gen_worker, args=(gemini_prompt,), daemon=True).start()

    def _gen_worker(self, prompt: str):
        _logger.info(f"[NB2] Single-still Gemini call — prompt_len={len(prompt)}")
        _logger.debug(f"[NB2] Full prompt:\n{prompt}")
        try:
            client   = self._init_nb2()
            response = client.models.generate_content(
                model="publishers/google/models/gemini-3.1-flash-image",
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_modalities=["IMAGE"],
                    image_config=types.ImageConfig(aspect_ratio="16:9"),
                ),
            )
            img_bytes = None
            if response.candidates and response.candidates[0].content.parts:
                for part in response.candidates[0].content.parts:
                    if part.inline_data and part.inline_data.data:
                        img_bytes = part.inline_data.data
                        break
            if img_bytes is None:
                raise ValueError("Model returned no image data.")
            _logger.info(f"[NB2] Image received — {len(img_bytes)} bytes")
            self.current_image_bytes = img_bytes
            self.after(0, lambda: self._show_generated(img_bytes))
        except Exception as exc:
            _logger.error(f"[NB2] Gemini error: {exc}")
            self.after(0, lambda: self._gen_error(str(exc)))

    def _show_generated(self, img_bytes: bytes):
        self._generating = False
        self.btn_generate.configure(state="normal", text="Generate Image")
        try:
            if self.selected_still:
                sid = self.selected_still["still_id"]
                self._pending_images[sid] = img_bytes
            self._img_data = Image.open(BytesIO(img_bytes))
            self.current_image_bytes = img_bytes
            self.btn_approve.configure(state="normal")
            self.after_idle(self._display_gen_image)
            self._populate_stills_list()
        except Exception as exc:
            messagebox.showerror("Display Error", str(exc))

    def _display_gen_image(self):
        if self._img_data is None:
            return
        # Read the stable FRAME dimensions, not the label (label size changes with image)
        fw = self._gen_frame.winfo_width()
        fh = self._gen_frame.winfo_height()
        w = max(fw - 20, 200)    # padx=10 each side
        h = max(fh - 126, 80)    # ~42px header + ~72px approve btn + 12px paddings
        img = self._img_data.copy()
        img.thumbnail((w, h), Image.LANCZOS)
        old_ref = getattr(self.gen_preview, "_ref", None)
        photo = ImageTk.PhotoImage(img)
        self.gen_preview.configure(image=photo, text="")
        self.gen_preview._ref = photo
        del old_ref  # safe to GC — tk.Label no longer holds the old name

    def _gen_error(self, msg: str):
        self._generating = False
        self.btn_generate.configure(state="normal", text="Generate Image")
        old_ref = getattr(self.gen_preview, "_ref", None)
        self.gen_preview.configure(
            image="",
            text="Generation failed — see error popup",
            fg="#A8372A", font=("Segoe UI", 13),
        )
        self.gen_preview._ref = None
        del old_ref
        messagebox.showerror("Generation Error", f"Image generation failed:\n{msg}")

    # ── Approval ──────────────────────────────────────────────────────────────

    def _approve_image(self):
        if not self.current_image_bytes or not self.selected_still:
            return
        self.output_dir.mkdir(parents=True, exist_ok=True)
        sid     = self.selected_still["still_id"]
        version = 1
        while (self.output_dir / f"{sid}_v{version}.png").exists():
            version += 1
        filename = f"{sid}_v{version}.png"
        out_path = self.output_dir / filename
        out_path.write_bytes(self.current_image_bytes)

        self.gen_state.setdefault("completed", {})[sid] = str(out_path)
        self._save_state()

        self._pending_images.pop(sid, None)
        self._img_data = None
        old_ref = getattr(self.gen_preview, "_ref", None)
        self.gen_preview.configure(
            image="", text="✅  Approved — select next still to continue.",
            fg=C["btn_green"], font=("Segoe UI", 13),
        )
        self.gen_preview._ref = None
        del old_ref
        self.btn_approve.configure(state="disabled")
        self.status_lbl.configure(text=f"  ✅ Saved: {filename}")
        self._chat_append("System:", f"✅ Image approved — saved as {filename}", "sys")
        self._populate_stills_list()
        messagebox.showinfo("Image Saved", f"Approved image saved to:\n{out_path}")

    # ── Settings persistence ──────────────────────────────────────────────────

    def _load_settings(self):
        if not SETTINGS_FILE.exists():
            return
        try:
            d = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
            if "art_style"     in d: self.var_style.set(d["art_style"])
            if "camera"        in d: self.var_camera.set(d["camera"])
            if "mood"          in d: self.var_mood.set(d["mood"])
            if "lighting"      in d: self.var_light.set(d["lighting"])
            if "color"         in d: self.var_color.set(d["color"])
            if "dof"           in d: self.var_dof.set(d["dof"])
            if "extra_notes"   in d:
                self.extra_notes.delete(0, "end")
                self.extra_notes.insert(0, d["extra_notes"])
            if "system_prompt" in d:
                self.sys_prompt_box.delete("0.0", "end")
                self.sys_prompt_box.insert("0.0", d["system_prompt"])
            if "output_dir"    in d:
                p = Path(d["output_dir"])
                self.output_dir = p
                self.out_dir_label.configure(text=str(p))
        except Exception:
            pass

    def _save_settings(self):
        try:
            d = {
                "art_style":     self.var_style.get(),
                "camera":        self.var_camera.get(),
                "mood":          self.var_mood.get(),
                "lighting":      self.var_light.get(),
                "color":         self.var_color.get(),
                "dof":           self.var_dof.get(),
                "extra_notes":   self.extra_notes.get().strip(),
                "system_prompt": self.sys_prompt_box.get("0.0", "end").strip(),
                "output_dir":    str(self.output_dir),
            }
            SETTINGS_FILE.write_text(json.dumps(d, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _on_close(self):
        self._save_settings()
        self.destroy()

    # ── Preview helpers ───────────────────────────────────────────────────────

    def _clear_preview(self):
        old = getattr(self.gen_preview, "_ref", None)
        self.gen_preview.configure(
            image="",
            text="Image will appear here after generation",
            fg=C["text_muted"], font=("Segoe UI", 13),
        )
        self.gen_preview._ref = None
        del old  # safe to GC now — widget no longer references the old image

    def _show_bulk_loading(self, sid: str):
        """Show loading state in preview panel if sid is the currently selected still."""
        if self.selected_still and self.selected_still["still_id"] == sid:
            old = getattr(self.gen_preview, "_ref", None)
            self.gen_preview.configure(
                image="",
                text="\n\n⏳\n\nGenerating image…",
                fg=C["accent"], font=("Segoe UI", 15, "bold"),
            )
            self.gen_preview._ref = None
            del old
            self.btn_approve.configure(state="disabled")

    # ── Bulk helpers (sidebar) ────────────────────────────────────────────────

    def _open_bulk_dialog(self):
        if not self.stills:
            messagebox.showwarning("No Stills", "Load a visual plan first.")
            return
        BulkGenerateDialog(self)

    def _bulk_approve_all(self):
        pending = dict(self._pending_images)
        if not pending:
            messagebox.showinfo("Nothing Pending", "No pending images to approve.")
            return
        n = len(pending)
        if not messagebox.askyesno(
            "Bulk Approve & Save All",
            f"Save all {n} pending generated images to disk?\n\nOutput: {self.output_dir}",
        ):
            return
        self.output_dir.mkdir(parents=True, exist_ok=True)
        for sid, img_bytes in pending.items():
            version = 1
            while (self.output_dir / f"{sid}_v{version}.png").exists():
                version += 1
            out_path = self.output_dir / f"{sid}_v{version}.png"
            out_path.write_bytes(img_bytes)
            self.gen_state.setdefault("completed", {})[sid] = str(out_path)
            self._pending_images.pop(sid, None)
        self._save_state()
        self._populate_stills_list()
        messagebox.showinfo("Saved", f"{n} images saved to:\n{self.output_dir}")


# ═════════════════════════════════════════════════════════════════════════════
class BulkGenerateDialog(ctk.CTkToplevel):
# ═════════════════════════════════════════════════════════════════════════════
    """
    Modal popup for bulk image generation with per-run settings, ETA timer,
    'Let AI Decide' per-setting option, system prompt, reference image, and
    resume-from-where-left-off support.
    """

    def __init__(self, parent: "ImageGenStudio"):
        super().__init__(parent)
        self._app       = parent
        self._running   = False
        self._cancel    = False
        self._start_ts  = 0.0
        self._gen_count = 0
        self.gpt_model_var = ctk.StringVar(value="gpt-4o")
        self._total     = 0
        self._errors: list[str] = []
        self._ref_b64: str | None = None

        self.title("Bulk Generate — Settings")
        self.geometry("760x980")
        self.resizable(True, True)
        self.configure(fg_color=C["app"])
        self.transient(parent)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build_ui()
        self._refresh_status()
        self._check_resume()

    # ── Live log helper ───────────────────────────────────────────────────────

    def _bulk_log(self, msg: str):
        """Append a timestamped line to the live log panel (thread-safe)."""
        ts   = time.strftime("%H:%M:%S")
        line = f"[{ts}]  {msg}\n"
        def _append():
            if hasattr(self, "log_box"):
                self.log_box.configure(state="normal")
                self.log_box.insert("end", line)
                self.log_box.see("end")
                self.log_box.configure(state="disabled")
        self.after(0, _append)

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        outer = ctk.CTkScrollableFrame(
            self, fg_color=C["app"],
            scrollbar_button_color=C["divider"],
            scrollbar_button_hover_color=C["accent"],
        )
        outer.pack(fill="both", expand=True)
        outer.grid_columnconfigure(0, weight=1)

        r = 0

        # Header
        ctk.CTkLabel(outer, text="Bulk Generate — Settings",
                     font=F(18, "bold"), text_color=C["accent"]).grid(
            row=r, column=0, sticky="w", padx=20, pady=(18, 6))
        r += 1

        # ── Image Settings ─────────────────────────────────────────────────
        sf = self._section(outer, "Image Settings")
        sf.grid(row=r, column=0, sticky="ew", padx=12, pady=(0, 8))
        sf.grid_columnconfigure((1, 3), weight=1)
        r += 1

        self.var_style  = ctk.StringVar(value=AI_DECIDE)
        self.var_camera = ctk.StringVar(value=AI_DECIDE)
        self.var_mood   = ctk.StringVar(value=AI_DECIDE)
        self.var_light  = ctk.StringVar(value=AI_DECIDE)
        self.var_color  = ctk.StringVar(value=AI_DECIDE)
        self.var_dof    = ctk.StringVar(value=AI_DECIDE)

        def lbl(parent, text):
            return ctk.CTkLabel(parent, text=text, font=F(13),
                                text_color=C["text_mid"])

        def row2(r, l1, v1, opts1, l2, v2, opts2):
            lbl(sf, l1).grid(row=r, column=0, padx=(14, 4), pady=5, sticky="w")
            make_combo(sf, [AI_DECIDE] + opts1, v1).grid(
                row=r, column=1, padx=(0, 12), pady=5, sticky="ew")
            lbl(sf, l2).grid(row=r, column=2, padx=(12, 4), pady=5, sticky="w")
            make_combo(sf, [AI_DECIDE] + opts2, v2).grid(
                row=r, column=3, padx=(0, 14), pady=5, sticky="ew")

        row2(1, "Art Style:", self.var_style, ART_STYLES,
                "Camera Angle:", self.var_camera, CAMERA_ANGLES)
        row2(2, "Mood:", self.var_mood, MOODS,
                "Lighting:", self.var_light, LIGHTING)
        row2(3, "Color Palette:", self.var_color, COLOR_PALETTES,
                "Depth of Field:", self.var_dof, DEPTH_OF_FIELD)

        lbl(sf, "Extra Notes:").grid(row=4, column=0, padx=(14, 4), pady=(4, 10), sticky="w")
        self.extra_notes = ctk.CTkEntry(
            sf, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1, font=F(13), height=34,
        )
        self.extra_notes.grid(row=4, column=1, columnspan=3,
                               padx=(0, 14), pady=(4, 10), sticky="ew")

        # ── Reference Image ────────────────────────────────────────────────
        rf = self._section(outer, "Reference Image")
        rf.grid(row=r, column=0, sticky="ew", padx=12, pady=(0, 8))
        rf.grid_columnconfigure(0, weight=1)
        r += 1

        br = ctk.CTkFrame(rf, fg_color="transparent")
        br.grid(row=1, column=0, padx=14, pady=(4, 4), sticky="ew")
        br.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(
            br, text="Browse…", width=120, height=32, corner_radius=6,
            font=F(13, "bold"), fg_color=C["btn_brown"], hover_color="#4E342E",
            command=self._browse_ref,
        ).grid(row=0, column=0, padx=(0, 10))
        self.ref_lbl = ctk.CTkLabel(
            br, text="No reference image", font=F(12),
            text_color=C["text_muted"], anchor="w")
        self.ref_lbl.grid(row=0, column=1, sticky="w")

        self.ref_preview = ctk.CTkLabel(
            rf, text="", height=0, fg_color="transparent")
        self.ref_preview.grid(row=2, column=0, padx=14, pady=0, sticky="ew")

        ctk.CTkLabel(rf, text="What to pick from this reference image:",
                     font=F(13), text_color=C["text_mid"]).grid(
            row=3, column=0, padx=14, pady=(6, 2), sticky="w")

        dlg_desc_row = ctk.CTkFrame(rf, fg_color="transparent")
        dlg_desc_row.grid(row=4, column=0, padx=14, pady=(0, 12), sticky="ew")
        dlg_desc_row.grid_columnconfigure(0, weight=1)

        self.ref_desc_entry = ctk.CTkEntry(
            dlg_desc_row, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1, font=F(13), height=34,
            placeholder_text="e.g. color grading, lighting style, mood, framing…",
        )
        self.ref_desc_entry.grid(row=0, column=0, sticky="ew")

        self.btn_extract = ctk.CTkButton(
            dlg_desc_row, text="Extract Settings",
            height=34, corner_radius=6, font=F(13, "bold"),
            fg_color=C["btn_red"], hover_color="#8B1A1A",
            command=self._extract_ref_settings,
        )
        self.btn_extract.grid(row=0, column=1, padx=(8, 0))

        # ── System Prompt ──────────────────────────────────────────────────
        pf = self._section(outer, "Style System Prompt  (applied to every still)")
        pf.grid(row=r, column=0, sticky="ew", padx=12, pady=(0, 8))
        pf.grid_columnconfigure(0, weight=1)
        r += 1

        gp_row = ctk.CTkFrame(pf, fg_color="transparent")
        gp_row.grid(row=1, column=0, padx=14, pady=(4, 6), sticky="ew")
        gp_row.grid_columnconfigure(2, weight=1)

        self.btn_gen_prompt = ctk.CTkButton(
            gp_row, text="Auto-Generate via GPT", height=32, corner_radius=6,
            font=F(13, "bold"), fg_color=C["btn_blue"], hover_color="#0D3D6E",
            command=self._auto_gen_prompt,
        )
        self.btn_gen_prompt.grid(row=0, column=0, padx=(0, 10))

        dlg_gm_combo = make_combo(gp_row, ALL_GPT_MODELS, self.gpt_model_var, width=170)
        dlg_gm_combo.grid(row=0, column=1)
        dlg_gm_combo.configure(command=lambda v: self._update_tier_lbl(v, self._dlg_tier_lbl))
        self._dlg_tier_lbl = ctk.CTkLabel(
            gp_row, text="● Higher Capacity", font=F(12), text_color="#27643B")
        self._dlg_tier_lbl.grid(row=0, column=2, padx=(8, 0), sticky="w")

        self.sys_prompt_box = ctk.CTkTextbox(
            pf, height=80, fg_color=C["input"], text_color=C["text"],
            border_color=C["divider"], border_width=1, font=F(13),
        )
        self.sys_prompt_box.grid(row=2, column=0, padx=14, pady=(0, 12), sticky="ew")
        self.sys_prompt_box.insert(
            "0.0",
            "Generate each image in a cohesive visual style that feels consistent "
            "across the entire video. Maintain the same lighting tone, color grading, "
            "and artistic approach throughout all stills.",
        )

        # ── Status + progress ──────────────────────────────────────────────
        self.lbl_status = ctk.CTkLabel(
            outer, text="", font=F(12), text_color=C["text_muted"], anchor="w")
        self.lbl_status.grid(row=r, column=0, padx=20, pady=(4, 2), sticky="ew")
        r += 1

        self.progress_bar = ctk.CTkProgressBar(
            outer, height=8, fg_color=C["divider"], progress_color=C["accent"])
        self.progress_bar.set(0)
        self.progress_bar.grid(row=r, column=0, padx=20, pady=(0, 2), sticky="ew")
        r += 1

        self.lbl_timer = ctk.CTkLabel(
            outer, text="", font=F(12), text_color=C["text_muted"], anchor="w")
        self.lbl_timer.grid(row=r, column=0, padx=20, pady=(0, 6), sticky="w")
        r += 1

        # ── Activity log panel ─────────────────────────────────────────────
        log_hdr = ctk.CTkFrame(outer, fg_color="transparent")
        log_hdr.grid(row=r, column=0, padx=20, pady=(0, 2), sticky="ew")
        log_hdr.grid_columnconfigure(0, weight=1)
        r += 1
        ctk.CTkLabel(log_hdr, text="Activity Log", font=F(12, "bold"),
                     text_color=C["text_mid"]).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(
            log_hdr, text="Clear", width=54, height=22, corner_radius=5,
            font=F(11), fg_color=C["chip_bg"], text_color=C["chip_text"],
            hover_color=C["divider"],
            command=lambda: (
                self.log_box.configure(state="normal"),
                self.log_box.delete("0.0", "end"),
                self.log_box.configure(state="disabled"),
            ),
        ).grid(row=0, column=1, sticky="e")

        self.log_box = ctk.CTkTextbox(
            outer, height=120, fg_color="#141414", text_color="#90EE90",
            border_color=C["divider"], border_width=1,
            font=ctk.CTkFont(family="Consolas", size=11),
            state="disabled",
        )
        self.log_box.grid(row=r, column=0, padx=12, pady=(0, 8), sticky="ew")
        r += 1

        # ── Action buttons ─────────────────────────────────────────────────
        brow = ctk.CTkFrame(outer, fg_color="transparent")
        brow.grid(row=r, column=0, padx=14, pady=(0, 24), sticky="ew")
        brow.grid_columnconfigure((0, 1), weight=1)
        r += 1

        self.btn_start = ctk.CTkButton(
            brow, text="Bulk Generate All", height=46, corner_radius=10,
            font=F(15, "bold"), fg_color=C["btn_green"], hover_color="#1B4D2E",
            command=self._toggle_generate,
        )
        self.btn_start.grid(row=0, column=0, padx=(0, 5), sticky="ew")

        self.btn_approve_all = ctk.CTkButton(
            brow, text="Bulk Approve & Save", height=46, corner_radius=10,
            font=F(15, "bold"), fg_color=C["btn_brown"], hover_color="#4E342E",
            command=self._bulk_approve_all,
        )
        self.btn_approve_all.grid(row=0, column=1, padx=(5, 0), sticky="ew")

        ctk.CTkButton(
            brow, text="↺", width=46, height=46, corner_radius=10,
            font=F(20, "bold"), fg_color=C["btn_red"], hover_color="#6B1A14",
            command=self._restart_from_scratch,
        ).grid(row=0, column=2, padx=(8, 0), sticky="e")

    def _section(self, parent, title: str) -> ctk.CTkFrame:
        f = ctk.CTkFrame(parent, fg_color=C["panel"], corner_radius=12,
                          border_width=1, border_color=C["divider"])
        ctk.CTkLabel(f, text=title, font=F(14, "bold"),
                     text_color=C["accent"]).grid(
            row=0, column=0, columnspan=4, padx=14, pady=(10, 2), sticky="w")
        return f

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _settings_block(self) -> str:
        def fmt(label: str, var: ctk.StringVar) -> str:
            v = var.get()
            if v == AI_DECIDE:
                return (f"{label}  [Choose the most visually appropriate "
                        f"{label.rstrip(':').lower()} for this scene based on its content]")
            return f"{label}  {v}"
        lines = [
            fmt("Art Style:",      self.var_style),
            fmt("Camera Angle:",   self.var_camera),
            fmt("Mood:",           self.var_mood),
            fmt("Lighting:",       self.var_light),
            fmt("Color Palette:",  self.var_color),
            fmt("Depth of Field:", self.var_dof),
        ]
        notes = self.extra_notes.get().strip()
        if notes:
            lines.append(f"Extra Notes:  {notes}")
        return "\n".join(lines)

    def _fmt_time(self, secs: float) -> str:
        m, s = divmod(int(secs), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"

    def _refresh_status(self):
        app       = self._app
        completed = app.gen_state.get("completed", {})
        pending   = app._pending_images
        total     = len(app.stills)
        done_n    = sum(1 for s in app.stills if s["still_id"] in completed)
        pend_n    = sum(1 for s in app.stills
                        if s["still_id"] in pending and s["still_id"] not in completed)
        todo_n    = total - done_n - pend_n
        self.lbl_status.configure(
            text=(f"{total} stills total  •  {done_n} approved  •  "
                  f"{pend_n} pending  •  {todo_n} not yet generated"),
            text_color=C["text_muted"],
        )

    # ── GPT tier helper ───────────────────────────────────────────────────────

    @staticmethod
    def _update_tier_lbl(model: str, lbl: ctk.CTkLabel):
        if model in HC_MODELS:
            lbl.configure(text="● Higher Capacity", text_color="#27643B")
        else:
            lbl.configure(text="● Higher Volume",   text_color="#B8860B")

    # ── Extract image settings from reference ─────────────────────────────────

    def _extract_ref_settings(self):
        if not self._ref_b64:
            messagebox.showwarning("No Reference Image",
                                   "Browse a reference image first.", parent=self)
            return
        if not OPENAI_API_KEY:
            messagebox.showerror("API Key Missing",
                                  "OPENAI_API_KEY not found in .env file.", parent=self)
            return
        self.btn_extract.configure(state="disabled", text="Extracting…")
        ref_b64  = self._ref_b64
        ref_desc = self.ref_desc_entry.get().strip()
        model    = self.gpt_model_var.get()
        threading.Thread(target=self._extract_settings_worker,
                          args=(ref_b64, ref_desc, model), daemon=True).start()

    def _extract_settings_worker(self, ref_b64: str, ref_desc: str, model: str):
        try:
            client = OpenAI(api_key=OPENAI_API_KEY)
            sys_msg, user_text = _build_extraction_prompt(ref_desc)
            msgs = [
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": [
                    {"type": "image_url",
                     "image_url": {"url": f"data:image/png;base64,{ref_b64}",
                                   "detail": "high"}},
                    {"type": "text", "text": user_text},
                ]},
            ]
            resp = _gpt_call(client, model, msgs, 1200)
            raw  = resp.choices[0].message.content.strip()
            m    = re.search(r"\{.*\}", raw, re.DOTALL)
            data = json.loads(m.group() if m else raw)
            self.after(0, lambda: self._apply_extracted_settings(data))
        except Exception as exc:
            self.after(0, lambda e=str(exc): (
                messagebox.showerror("Extraction Error", e, parent=self),
                self.btn_extract.configure(state="normal", text="Extract Settings"),
            ))

    def _apply_extracted_settings(self, data: dict):
        def try_set(var, key, options):
            val = data.get(key, "")
            if val in options:
                var.set(val)

        try_set(self.var_style,  "art_style",      ART_STYLES)
        try_set(self.var_camera, "camera_angle",   CAMERA_ANGLES)
        try_set(self.var_mood,   "mood",           MOODS)
        try_set(self.var_light,  "lighting",       LIGHTING)
        try_set(self.var_color,  "color_palette",  COLOR_PALETTES)
        try_set(self.var_dof,    "depth_of_field", DEPTH_OF_FIELD)

        notes = data.get("extra_notes", "").strip()
        if notes:
            self.extra_notes.delete(0, "end")
            self.extra_notes.insert(0, notes)

        style_p = data.get("style_prompt", "").strip()
        subject = data.get("main_subject", "").strip()

        # Combine style directive + subject lock so both GPT and Gemini see them
        full_sys = style_p
        if subject:
            full_sys = f"{style_p}\n\n{subject}" if style_p else subject

        if full_sys:
            self.sys_prompt_box.delete("0.0", "end")
            self.sys_prompt_box.insert("0.0", full_sys)

        _logger.info(
            f"[BULK EXTRACT→UI] style_len={len(style_p)} | "
            f"subject={subject[:100]!r}"
        )
        self._bulk_log(f"Extracted — art={data.get('art_style')} | {subject[:80]!r}")

        # Sync to main frontend
        self._app._apply_bulk_settings({
            "art_style":     data.get("art_style", ""),
            "camera":        data.get("camera_angle", ""),
            "mood":          data.get("mood", ""),
            "lighting":      data.get("lighting", ""),
            "color":         data.get("color_palette", ""),
            "dof":           data.get("depth_of_field", ""),
            "extra_notes":   notes,
            "system_prompt": full_sys,
        })

        self.btn_extract.configure(state="normal", text="Extract Settings")
        messagebox.showinfo(
            "Settings Extracted",
            "Dropdowns, Extra Notes, and Style Prompt applied to bulk dialog and main frontend.\n"
            + (f"Subject lock: {subject[:80]}…" if subject else ""),
            parent=self,
        )

    # ── Reference image ───────────────────────────────────────────────────────

    def _browse_ref(self):
        path = filedialog.askopenfilename(
            parent=self, title="Choose reference image",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.webp *.bmp")],
        )
        if not path:
            return
        try:
            img = Image.open(path)
            buf = BytesIO()
            img.save(buf, format="PNG")
            self._ref_b64 = base64.b64encode(buf.getvalue()).decode()

            thumb = img.copy()
            thumb.thumbnail((700, 160), Image.LANCZOS)
            photo = ImageTk.PhotoImage(thumb)
            self.ref_preview.configure(image=photo, text="", height=thumb.height + 8)
            self.ref_preview._ref = photo
            self.ref_lbl.configure(text="")  # hide filename — thumbnail is enough
        except Exception as exc:
            messagebox.showerror("Image Error", str(exc), parent=self)

    # ── Auto-generate system prompt ───────────────────────────────────────────

    def _auto_gen_prompt(self):
        if not OPENAI_API_KEY:
            messagebox.showerror(
                "API Key Missing",
                "OPENAI_API_KEY not found in .env file.", parent=self)
            return
        if not self._app.stills:
            messagebox.showwarning("No Stills", "Load a visual plan first.", parent=self)
            return
        self.btn_gen_prompt.configure(state="disabled", text="Generating…")
        vos      = "\n".join(f"- {s['voiceover']}" for s in self._app.stills[:12])
        settings = self._settings_block()
        ref_b64  = self._ref_b64
        ref_desc = self.ref_desc_entry.get().strip()
        model    = self.gpt_model_var.get()
        threading.Thread(target=self._prompt_worker,
                          args=(vos, settings, ref_b64, ref_desc, model), daemon=True).start()

    def _prompt_worker(self, vos: str, settings: str, ref_b64: str | None = None,
                       ref_desc: str = "", model: str = "gpt-4o"):
        _logger.info(f"[GPT] Auto-generate style directive | model={model} | ref_desc={ref_desc!r}")
        try:
            client = OpenAI(api_key=OPENAI_API_KEY)
            user_parts: list = []
            if ref_b64:
                user_parts.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{ref_b64}", "detail": "high"},
                })
                if ref_desc:
                    user_parts.append({"type": "text", "text": (
                        f"CRITICAL FOCUS: From this reference image, extract SPECIFICALLY: {ref_desc}. "
                        f"Your style directive MUST describe the '{ref_desc}' approach in explicit, "
                        f"technical detail. This is the primary purpose — do not describe other aspects "
                        f"unless they directly relate to '{ref_desc}'."
                    )})
                else:
                    user_parts.append({"type": "text",
                                       "text": "Analyze this reference image's art style and rendering technique."})
            focus_note = (
                f"IMPORTANT: The style directive must explicitly describe the '{ref_desc}' approach "
                f"extracted from the reference image.\n"
                if ref_desc else ""
            )
            user_parts.append({"type": "text", "text": (
                f"I am generating AI images for a video. Voiceover lines (up to 12 stills):\n{vos}\n\n"
                f"Image settings:\n{settings}\n\n"
                f"Write a visual style directive (2-3 sentences) that will be prepended to "
                f"every image generation request.\n"
                f"{focus_note}"
                f"Rules:\n"
                f"- Describe HOW the images should look: art style, rendering technique, "
                f"line work, shading method, texture approach, and lighting.\n"
                f"- Be EXPLICIT: say 'Use bold black outlines with flat cel-shading' "
                f"— NOT 'ensure consistency' or 'capture the reference style'.\n"
                f"- Do NOT specify colours for the main subject (cat, person, animal) — "
                f"only describe the rendering technique and overall scene atmosphere.\n"
                f"- Do NOT start with 'Create a series of' or 'Generate images of'.\n"
                f"- Do NOT reference 'the reference image' in your output — "
                f"incorporate its qualities directly as style instructions.\n"
                f"- Return ONLY the directive text, no labels, no preamble."
            )})
            resp = _gpt_call(client, model,
                             [{"role": "user", "content": user_parts}], 350)
            text = resp.choices[0].message.content.strip()
            _logger.info(f"[GPT] Style directive result: {text[:300]!r}")
            self.after(0, lambda: self._apply_sys_prompt(text))
        except Exception as exc:
            _logger.error(f"[GPT] Style directive error: {exc}")
            self.after(0, lambda e=str(exc): (
                messagebox.showerror("GPT Error", e, parent=self),
                self.btn_gen_prompt.configure(
                    state="normal", text="Auto-Generate via GPT"),
            ))

    def _apply_sys_prompt(self, text: str):
        # Preserve any SUBJECT STYLE section from the existing system prompt
        existing = self.sys_prompt_box.get("0.0", "end").strip()
        preserved = ""
        if existing:
            lines = existing.split("\n")
            for i, line in enumerate(lines):
                if line.strip().upper().startswith("SUBJECT STYLE") or \
                   line.strip().upper().startswith("SUBJECT LOCK"):
                    preserved = "\n".join(lines[i:]).strip()
                    break
        new_content = f"{text}\n\n{preserved}" if preserved else text
        self.sys_prompt_box.delete("0.0", "end")
        self.sys_prompt_box.insert("0.0", new_content)
        self.btn_gen_prompt.configure(state="normal", text="Auto-Generate via GPT")

    # ── Resume check ──────────────────────────────────────────────────────────

    def _check_resume(self):
        app       = self._app
        completed = app.gen_state.get("completed", {})
        pending   = [s for s in app.stills
                     if s["still_id"] in app._pending_images
                     and s["still_id"] not in completed]
        if not pending:
            return
        ans = messagebox.askyesnocancel(
            "Resume Previous Run?",
            f"{len(pending)} stills from a previous bulk run are pending approval.\n\n"
            f"Yes  →  Skip already-generated stills (resume)\n"
            f"No   →  Regenerate everything (start over)\n"
            f"Cancel  →  Do nothing",
            parent=self,
        )
        if ans is False:
            app._pending_images.clear()
            app._populate_stills_list()
            self._refresh_status()

    # ── Timer ─────────────────────────────────────────────────────────────────

    def _tick_timer(self):
        if not self._running:
            return
        elapsed = time.time() - self._start_ts
        done    = self._gen_count
        elapsed_str = self._fmt_time(elapsed)
        if done > 0:
            eta_sec = (self._total - done) * (elapsed / done)
            eta_str = self._fmt_time(eta_sec)
        else:
            eta_str = "—"
        self.lbl_timer.configure(
            text=f"Elapsed: {elapsed_str}   |   ETA: {eta_str}")
        self.after(1000, self._tick_timer)

    # ── Generation ────────────────────────────────────────────────────────────

    def _toggle_generate(self):
        if self._running:
            self._cancel = True
            self.btn_start.configure(text="Stopping…", state="disabled")
            return

        app       = self._app
        completed = app.gen_state.get("completed", {})
        targets   = [s for s in app.stills
                     if s["still_id"] not in completed
                     and s["still_id"] not in app._pending_images]

        if not targets:
            pending_n = sum(1 for s in app.stills
                            if s["still_id"] not in completed
                            and s["still_id"] in app._pending_images)
            if pending_n:
                messagebox.showinfo(
                    "All Generated",
                    f"All stills are already generated ({pending_n} pending approval).\n"
                    f"Click 'Bulk Approve & Save' to save them to disk.",
                    parent=self,
                )
            else:
                messagebox.showinfo("All Done",
                                    "All stills have been approved.", parent=self)
            return

        # Snapshot all tkinter values in main thread before handing off to thread
        sys_txt          = self.sys_prompt_box.get("0.0", "end").strip()
        ref_b64_snapshot = self._ref_b64
        ref_desc_snap    = self.ref_desc_entry.get().strip()
        gpt_model_snap   = self.gpt_model_var.get()

        def _concrete(var):
            v = var.get()
            return v if v != AI_DECIDE else None

        bulk_settings = {
            "art_style":     _concrete(self.var_style),
            "camera":        _concrete(self.var_camera),
            "mood":          _concrete(self.var_mood),
            "lighting":      _concrete(self.var_light),
            "color":         _concrete(self.var_color),
            "dof":           _concrete(self.var_dof),
            "extra_notes":   self.extra_notes.get().strip(),
            "system_prompt": sys_txt,
        }

        self._running   = True
        self._cancel    = False
        self._gen_count = 0
        self._total     = len(targets)
        self._errors    = []
        self._start_ts  = time.time()

        self.btn_start.configure(
            text="Stop Generation", state="normal",
            fg_color=C["btn_red"], hover_color="#6B1A14",
        )
        self.progress_bar.set(0)
        self.lbl_timer.configure(text="Elapsed: 00:00   |   ETA: —")
        self._tick_timer()

        threading.Thread(
            target=self._bulk_worker,
            args=(targets, bulk_settings, sys_txt,
                  ref_b64_snapshot, ref_desc_snap, gpt_model_snap),
            daemon=True,
        ).start()

    def _bulk_worker(self, targets: list, bulk_settings: dict,
                     sys_txt: str,
                     ref_b64: str | None = None, ref_desc: str = "",
                     gpt_model: str = "gpt-4o"):
        _logger.info(
            f"[BULK] Worker started — {len(targets)} still(s) | "
            f"model={gpt_model} | ref_desc={ref_desc!r} | "
            f"sys_txt_len={len(sys_txt)} | has_ref={'yes' if ref_b64 else 'no'}"
        )
        self._bulk_log(f"Starting bulk gen — {len(targets)} still(s) | GPT: {gpt_model}")
        try:
            gemini_client = self._app._init_nb2()
            _logger.info("[BULK] Gemini (NB2) client initialised OK")
            self._bulk_log("Gemini (NB2) client connected OK")
        except Exception as exc:
            _logger.error(f"[BULK] Gemini init failed: {exc}")
            self.after(0, lambda e=str(exc): (
                messagebox.showerror("Gemini Error",
                                     f"Failed to connect to Gemini:\n{e}", parent=self),
                self._bulk_done(),
            ))
            return

        gpt_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
        if not gpt_client:
            _logger.warning("[BULK] No OpenAI key — skipping GPT prompt writing")
            self._bulk_log("WARNING: No OpenAI key — using fallback prompts")

        # Ref image thumbnail bytes for main frontend display (decoded once)
        ref_thumb_b64 = ref_b64  # keep as b64 string; decoded in _on_gen callback

        for still in targets:
            if self._cancel:
                _logger.info("[BULK] Cancelled by user")
                self._bulk_log("Cancelled by user.")
                break
            sid = still["still_id"]
            n   = self._gen_count
            vo  = str(still.get("voiceover", ""))

            _logger.info(f"[BULK][{sid}] ── Processing still — vo={vo[:80]!r}")
            self._bulk_log(f"[{sid.upper()}] Starting — vo: {vo[:70]!r}")

            # Show loading state in main window preview if this still is selected
            self.after(0, lambda s=sid: self._app._show_bulk_loading(s))

            # ── Step 0: Resolve "Let AI Decide" fields for this still ────────────
            per_still_settings = dict(bulk_settings)
            ai_decide_fields = {k: v for k, v in per_still_settings.items()
                                if v is None and k in FIELD_OPTIONS_MAP}

            if ai_decide_fields:
                _logger.info(f"[BULK][{sid}] AI-Decide fields: {list(ai_decide_fields.keys())}")
                self._bulk_log(f"[{sid.upper()}] AI-Decide: resolving {list(ai_decide_fields.keys())}")

            if ai_decide_fields and gpt_client:
                field_lines = "\n".join(
                    f'"{k}": one of {json.dumps([o for o in FIELD_OPTIONS_MAP[k] if "Custom" not in o])}'
                    for k in ai_decide_fields
                )
                resolve_msg = (
                    f"Scene voiceover: \"{still['voiceover']}\"\n\n"
                    f"Choose the single best value for each field that fits this specific scene:\n"
                    f"{field_lines}\n\n"
                    f"Return ONLY valid JSON with those keys."
                )
                try:
                    r = _gpt_call(gpt_client, gpt_model,
                                  [{"role": "user", "content": resolve_msg}], 200)
                    raw = r.choices[0].message.content.strip()
                    m = re.search(r"\{.*\}", raw, re.DOTALL)
                    resolved = json.loads(m.group() if m else raw)
                    for k, options in FIELD_OPTIONS_MAP.items():
                        if k in ai_decide_fields and resolved.get(k) in options:
                            per_still_settings[k] = resolved[k]
                    resolved_log = {k: per_still_settings[k] for k in ai_decide_fields}
                    _logger.info(f"[BULK][{sid}] AI-Decide resolved: {resolved_log}")
                    self._bulk_log(f"[{sid.upper()}] AI-Decide → {resolved_log}")
                except Exception as exc:
                    _logger.warning(f"[BULK][{sid}] AI-Decide GPT call failed: {exc}")
                    self._bulk_log(f"[{sid.upper()}] AI-Decide GPT failed — using random fallback")
                # Fallback for any still-unresolved fields: pick randomly (exclude Custom)
                for k, options in FIELD_OPTIONS_MAP.items():
                    if per_still_settings.get(k) is None:
                        safe_opts = [o for o in options if "Custom" not in o]
                        per_still_settings[k] = random.choice(safe_opts)
                        _logger.info(f"[BULK][{sid}] Random fallback for '{k}': {per_still_settings[k]}")

            # Build a concrete settings block using per-still resolved values
            label_map = [
                ("art_style", "Art Style"),
                ("camera",    "Camera Angle"),
                ("mood",      "Mood"),
                ("lighting",  "Lighting"),
                ("color",     "Color Palette"),
                ("dof",       "Depth of Field"),
            ]
            concrete_block_lines = [
                f"{lbl}: {per_still_settings[k]}"
                for k, lbl in label_map if per_still_settings.get(k)
            ]
            notes = per_still_settings.get("extra_notes", "")
            if notes:
                concrete_block_lines.append(f"Extra Notes: {notes}")
            concrete_settings_block = "\n".join(concrete_block_lines)
            _logger.info(f"[BULK][{sid}] Resolved settings:\n{concrete_settings_block}")

            # ── Phase 1: GPT writes a unique creative prompt for this scene ──────
            creative_prompt = None
            if gpt_client:
                self.after(0, lambda sid=sid, n=n: self.lbl_status.configure(
                    text=f"Writing prompt for {sid.upper()}…  ({n} / {self._total} done)",
                    text_color=C["accent"],
                ))
                self._bulk_log(f"[{sid.upper()}] Phase 1 — GPT writing creative prompt…")
                try:
                    # ── DO NOT send ref_b64 here ────────────────────────────────────
                    # Sending the reference image to GPT for creative prompts causes it
                    # to copy the reference CONTENT/COMPOSITION (e.g. "man holding cat")
                    # into every scene. Style information is already captured in sys_txt.
                    # ref_b64 is used ONLY for style extraction, never for prompt writing.

                    # Extract the full SUBJECT STYLE paragraph from sys_txt
                    subject_style = ""
                    if sys_txt:
                        lines = sys_txt.split("\n")
                        subject_start = -1
                        for i, line in enumerate(lines):
                            if line.strip().upper().startswith("SUBJECT STYLE") or \
                               line.strip().upper().startswith("SUBJECT LOCK"):
                                subject_start = i
                                break
                        if subject_start >= 0:
                            subject_lines = []
                            for line in lines[subject_start:]:
                                s = line.strip()
                                if not s and subject_lines:
                                    break
                                if s:
                                    subject_lines.append(s)
                            subject_style = " ".join(subject_lines)

                    ref_desc_rule = (
                        f"- STYLE NOTE: The system prompt captures the rendering style of "
                        f"'{ref_desc}' from a reference. Apply those extracted line quality, "
                        f"colours, and texture instructions. "
                        f"Do NOT copy any composition, poses, or scene content from the reference.\n"
                        if ref_desc else ""
                    )
                    subject_rule = (
                        f"- SUBJECT APPEARANCE (HIGHEST PRIORITY): If the scene includes the "
                        f"main subject (cat or animal), you MUST describe it using EXACTLY these "
                        f"colours and details — override any palette colours listed above for "
                        f"the subject itself:\n"
                        f"  {subject_style}\n"
                        if subject_style else ""
                    )
                    # Build mandatory settings list so GPT cannot skip any of them
                    mandatory = []
                    _sm = per_still_settings
                    if _sm.get("art_style"):
                        mandatory.append(f'Art Style "{_sm["art_style"]}" — name it explicitly')
                    if _sm.get("camera"):
                        mandatory.append(f'Camera Angle "{_sm["camera"]}" — describe the framing')
                    if _sm.get("mood"):
                        mandatory.append(f'Mood "{_sm["mood"]}" — reflect it in atmosphere')
                    if _sm.get("lighting"):
                        mandatory.append(f'Lighting "{_sm["lighting"]}" — describe the light')
                    if _sm.get("color"):
                        mandatory.append(
                            f'Color Palette "{_sm["color"]}" — apply to the SCENE/BACKGROUND '
                            f'atmosphere, NOT to the main subject (subject colours come from '
                            f'SUBJECT APPEARANCE above)'
                        )
                    if _sm.get("dof"):
                        mandatory.append(f'Depth of Field "{_sm["dof"]}"')
                    mandatory_block = "\n".join(f"  • {m}" for m in mandatory)

                    gpt_parts = [{"type": "text", "text": (
                        f"Write a vivid image-generation prompt for this specific scene.\n\n"
                        f"Voiceover: \"{still['voiceover']}\"\n\n"
                        f"{subject_rule}"
                        f"MANDATORY — your prompt MUST explicitly include ALL of these:\n"
                        f"{mandatory_block}\n\n"
                        f"Rules:\n"
                        f"- VISUALLY INTERPRET the voiceover — describe what the viewer SEES. "
                        f"Do NOT copy voiceover text verbatim.\n"
                        f"- Derive composition from the voiceover ONLY — do NOT borrow poses or "
                        f"scene layout from any reference image.\n"
                        f"- Be specific: subject, background, colours, atmosphere.\n"
                        f"{ref_desc_rule}"
                        f"- 3-5 sentences — no labels, no preamble, just the prompt text."
                    )}]
                    gpt_msgs = []
                    if sys_txt:
                        gpt_msgs.append({"role": "system", "content": sys_txt})
                    gpt_msgs.append({"role": "user", "content": gpt_parts})

                    gpt_resp = _gpt_call(gpt_client, gpt_model, gpt_msgs, 400)
                    creative_prompt = gpt_resp.choices[0].message.content.strip()
                    _logger.info(f"[BULK][{sid}] GPT creative prompt: {creative_prompt[:250]!r}")
                    self._bulk_log(f"[{sid.upper()}] GPT prompt OK ({len(creative_prompt)} chars)")
                except Exception as gpt_exc:
                    _logger.error(f"[BULK][{sid}] GPT prompt error: {gpt_exc}")
                    self._bulk_log(f"[{sid.upper()}] GPT error — using fallback prompt")
                    self._errors.append(f"{sid} (GPT): {str(gpt_exc)[:80]}")

            # Fallback if GPT unavailable or failed
            if not creative_prompt:
                creative_prompt = f"{concrete_settings_block}\n\nScene: {still['voiceover']}"
                _logger.info(f"[BULK][{sid}] Using fallback prompt (no GPT)")
                self._bulk_log(f"[{sid.upper()}] Using fallback prompt (no GPT)")

            # ── Assemble final Gemini prompt ──────────────────────────────────────
            # Structure:
            #   1. Style directive + SUBJECT STYLE (from sys_txt)
            #   2. Optional ref_desc rendering-style focus note
            #   3. Scene description (from GPT creative_prompt)
            #   4. Guaranteed settings footer — Gemini ALWAYS sees every setting
            #      even if GPT failed to mention one of them
            parts = []
            if sys_txt:
                parts.append(sys_txt)
            if ref_desc:
                parts.append(
                    f"RENDERING STYLE FOCUS (not composition): Apply the drawing/rendering "
                    f"style of '{ref_desc}' from the extracted reference — meaning line quality, "
                    f"colour values, shading technique, and texture. "
                    f"Do NOT replicate any composition, pose, or scene layout from the reference image."
                )
            parts.append(creative_prompt)
            if concrete_settings_block:
                parts.append(
                    f"APPLY THESE STYLE SETTINGS EXACTLY:\n{concrete_settings_block}"
                )
            gemini_prompt = "\n\n".join(parts)
            _logger.info(
                f"[BULK][{sid}] Gemini prompt assembled — "
                f"sys_txt_len={len(sys_txt)} | creative_len={len(creative_prompt)} | "
                f"total_len={len(gemini_prompt)}"
            )
            _logger.debug(f"[BULK][{sid}] Full Gemini prompt:\n{gemini_prompt}")

            # ── Phase 2: Gemini generates the image (text prompt only, no image) ─
            self.after(0, lambda sid=sid, n=n: self.lbl_status.configure(
                text=f"Generating image for {sid.upper()}…  ({n} / {self._total} done)",
                text_color=C["accent"],
            ))
            self._bulk_log(f"[{sid.upper()}] Phase 2 — sending to Gemini (NB2)…")

            attempt = 0
            while not self._cancel:
                attempt += 1
                try:
                    response = gemini_client.models.generate_content(
                        model="publishers/google/models/gemini-3.1-flash-image",
                        contents=gemini_prompt,
                        config=types.GenerateContentConfig(
                            response_modalities=["IMAGE"],
                            image_config=types.ImageConfig(aspect_ratio="16:9"),
                        ),
                    )
                    img_bytes = None
                    if response.candidates and response.candidates[0].content.parts:
                        for part in response.candidates[0].content.parts:
                            if part.inline_data and part.inline_data.data:
                                img_bytes = part.inline_data.data
                                break

                    if img_bytes:
                        _logger.info(f"[BULK][{sid}] Gemini returned image — {len(img_bytes)} bytes")
                        self._bulk_log(f"[{sid.upper()}] Image received — {len(img_bytes)//1024} KB ✓")
                        self._app._pending_images[sid] = img_bytes
                        self._gen_count += 1

                        def _on_gen(sid=sid, ib=img_bytes, p=creative_prompt,
                                    bs=per_still_settings, rb=ref_thumb_b64):
                            app = self._app
                            app._still_states[sid] = {
                                "history": [{"role": "assistant", "content": p}],
                                "log":     [("gpt", "Bulk-generated prompt:", p)],
                                "prompt":  p,
                                "bulk_settings": bs,
                            }
                            app._populate_stills_list()
                            # Mirror ref image to main frontend on first successful gen
                            if rb and not app.ref_image_b64:
                                try:
                                    ref_img = Image.open(BytesIO(base64.b64decode(rb)))
                                    thumb = ref_img.copy()
                                    thumb.thumbnail((308, 192), Image.LANCZOS)
                                    photo_ref = ImageTk.PhotoImage(thumb)
                                    app.ref_label.configure(image=photo_ref, text="")
                                    app.ref_label._ref = photo_ref
                                    app.ref_image_b64 = rb
                                except Exception:
                                    pass
                            if app.selected_still and app.selected_still["still_id"] == sid:
                                app.current_image_bytes = ib
                                app._img_data = Image.open(BytesIO(ib))
                                app.btn_approve.configure(state="normal")
                                app.after_idle(app._display_gen_image)
                                app._apply_bulk_settings(bs)
                                app.prompt_editor.delete("0.0", "end")
                                app.prompt_editor.insert("0.0", p)
                                app._chat_append("Bulk-generated prompt:", p, "gpt")
                                app.chat_history = [{"role": "assistant", "content": p}]
                            self.progress_bar.set(self._gen_count / self._total)

                        self.after(0, _on_gen)
                    else:
                        _logger.warning(f"[BULK][{sid}] Gemini returned no image data")
                        self._bulk_log(f"[{sid.upper()}] WARNING: Gemini returned no image")
                        self._errors.append(f"{sid}: Gemini returned no image")
                    time.sleep(2)
                    break

                except Exception as exc:
                    err_str = str(exc)
                    if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str:
                        _logger.warning(f"[BULK][{sid}] Rate limited (attempt {attempt}) — retrying…")
                        self._bulk_log(f"[{sid.upper()}] Rate limited — retrying in 10 s…")
                        for countdown in range(10, 0, -1):
                            if self._cancel:
                                break
                            self.after(0, lambda s=sid, c=countdown, a=attempt: self.lbl_status.configure(
                                text=f"Rate limited on {s.upper()} (attempt {a}) — retrying in {c}s…",
                                text_color=C["accent"],
                            ))
                            time.sleep(1)
                    else:
                        _logger.error(f"[BULK][{sid}] Gemini error: {err_str}")
                        self._bulk_log(f"[{sid.upper()}] ERROR: {err_str[:100]}")
                        self._errors.append(f"{sid}: {err_str[:100]}")
                        break

        _logger.info(f"[BULK] Worker finished — {self._gen_count}/{self._total} generated | errors={len(self._errors)}")
        self._bulk_log(f"Done — {self._gen_count}/{self._total} generated | {len(self._errors)} error(s)")
        self.after(0, self._bulk_done)

    def _bulk_done(self):
        self._running = False
        self._cancel  = False
        done  = self._gen_count
        errs  = len(self._errors)

        self.btn_start.configure(
            text="Bulk Generate All", state="normal",
            fg_color=C["btn_green"], hover_color="#1B4D2E",
        )
        self.lbl_timer.configure(text="")
        self._refresh_status()
        self._app._populate_stills_list()

        msg = f"Done — {done} / {self._total} generated"
        if errs:
            msg += f"  ({errs} failed)"
        self.lbl_status.configure(
            text=msg,
            text_color=C["btn_red"] if errs else C["btn_green"],
        )
        self.progress_bar.set(done / self._total if self._total else 0)

        if errs and messagebox.askyesno(
            "Generation Errors",
            f"{errs} still(s) failed. Show details?",
            parent=self,
        ):
            messagebox.showinfo(
                "Error Details", "\n".join(self._errors), parent=self)

    # ── Restart from scratch ──────────────────────────────────────────────────

    def _restart_from_scratch(self):
        if self._running:
            messagebox.showwarning("Running",
                                   "Stop generation first before restarting.", parent=self)
            return
        if not messagebox.askyesno(
            "Restart from Scratch",
            "Clear all pending (not-yet-approved) generated images and start fresh?\n\n"
            "Already approved & saved images will NOT be affected.",
            parent=self,
        ):
            return
        app = self._app
        app._pending_images.clear()
        for sid in list(app._still_states.keys()):
            if "bulk_settings" in app._still_states[sid]:
                del app._still_states[sid]
        self._gen_count = 0
        self.progress_bar.set(0)
        self._refresh_status()
        app._populate_stills_list()
        self.lbl_status.configure(
            text="Pending images cleared — ready for a fresh run.",
            text_color=C["btn_green"],
        )

    # ── Bulk approve ──────────────────────────────────────────────────────────

    def _bulk_approve_all(self):
        pending = {sid: ib for sid, ib in self._app._pending_images.items()
                   if sid not in self._app.gen_state.get("completed", {})}
        if not pending:
            messagebox.showinfo("Nothing Pending",
                                "No pending images to approve.", parent=self)
            return
        n = len(pending)
        if not messagebox.askyesno(
            "Bulk Approve & Save",
            f"Save all {n} pending generated images to disk?\n\nOutput: {self._app.output_dir}",
            parent=self,
        ):
            return
        self._app.output_dir.mkdir(parents=True, exist_ok=True)
        for sid, img_bytes in pending.items():
            version = 1
            while (self._app.output_dir / f"{sid}_v{version}.png").exists():
                version += 1
            out_path = self._app.output_dir / f"{sid}_v{version}.png"
            out_path.write_bytes(img_bytes)
            self._app.gen_state.setdefault("completed", {})[sid] = str(out_path)
            self._app._pending_images.pop(sid, None)
        self._app._save_state()
        self._app._populate_stills_list()
        self._refresh_status()
        messagebox.showinfo(
            "Saved", f"{n} images saved to:\n{self._app.output_dir}", parent=self)

    # ── Close ─────────────────────────────────────────────────────────────────

    def _on_close(self):
        if self._running:
            if not messagebox.askyesno(
                "Stop Generation?",
                "Bulk generation is running. Stop and close?",
                parent=self,
            ):
                return
            self._cancel = True
        self.destroy()


# ═════════════════════════════════════════════════════════════════════════════
def main():
    app = ImageGenStudio()
    app.mainloop()

if __name__ == "__main__":
    main()
